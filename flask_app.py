"""
Copyright © 2023 Philippe Mourey

This script provides CRUD features inside a Flask application for job's research follow-up and contact recruiters at monthly basis using a scheduler

"""
from __future__ import annotations

import atexit
import locale
import logging
import os

from flask import (Flask, flash, redirect, render_template, request, session,
                   url_for)

# Ensure NO_PROXY is set (some environments use lowercase 'no_proxy') so requests bypasses
# the corporate proxy for France Travail domains when the variable was exported locally.
if not os.environ.get('NO_PROXY') and os.environ.get('no_proxy'):
    os.environ['NO_PROXY'] = os.environ.get('no_proxy')
os.environ.setdefault('NO_PROXY', 'entreprise.francetravail.fr,api.francetravail.io')
import re
from datetime import datetime, timedelta
import json as _json
from functools import wraps
from io import BytesIO
from logging import DEBUG, basicConfig
from pathlib import Path
from socket import socket
from time import sleep
from typing import Match, Optional

import pytz
from _socket import gethostbyname
from apscheduler.schedulers.background import BackgroundScheduler
from dateutil.relativedelta import relativedelta
from flask import (Flask, flash, jsonify, redirect, render_template, request,
                   send_file, session, url_for)
from flask_debugtoolbar import DebugToolbarExtension
from flask_login import (LoginManager, current_user, login_required,
                         login_user, logout_user)
from flask_sqlalchemy import SQLAlchemy
from itsdangerous import Serializer, URLSafeSerializer
from pytz import timezone
from sqlalchemy import DateTime, case, desc, func
from user_agents import parse
from user_agents.parsers import UserAgent
from validators import url
from werkzeug.security import generate_password_hash  # , check_password_hash

from Controller import (check, check_password_and_upgrade,
                        get_session_by_login, get_user_by_id,
                        handle_file_upload, send_confirmation_email,
                        send_password_recovery_email)
from Model import AppSetting, FtSearch, Job, Session, User, db
from tools.cv_tools import (build_cv_pdf_filename,
                            generate_tailored_cv_pdf_bytes,
                            get_ai_cover_letter_text, get_ai_cv_suggestions,
                            load_cv_data)
from tools.merge_cvs import merge_cv_jsons
from tools.import_cv import parse_json_cv, fetch_linkedin_profile
from tools.document_tools import (build_cover_letter_pdf_filename,
                                  generate_cover_letter_pdf_bytes,
                                  resolve_cover_letter_template_path)
from tools.france_travail import (ATS_KEYWORDS_PROFILE, CONTRACT_TYPES,
                                  DEPARTMENTS, WORK_MODES, search_auto_from_cv,
                                  search_offers)
from tools.send_emails import send_email

app = Flask(__name__, static_folder='static', static_url_path='/static')
# Set the environment (development, production, etc.)
# Replace 'development' with the appropriate value for your environment
app.config.from_object('config.Config')
# app.config.from_pyfile(config_filename)

db.init_app(app)

toolbar = DebugToolbarExtension(app)

locale.setlocale(locale.LC_TIME, 'fr_FR')
basicConfig(level=DEBUG)

login_manager = LoginManager(app)
login_manager.login_view = 'login'

# Set the secret key to some random bytes. Keep this really secret!
app.secret_key = b'_5#y2L"F4Q8z\n\xec]/'
# app.config['SECRET_KEY'] = 'fifa2022'
# app.config['SESSION_TYPE'] = 'filesystem'

UPLOAD_FOLDER = os.path.join('static', 'uploads')
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)

""" decorators """


def is_connected(func):
    @wraps(func)
    def wrapper(*args, **kwargs):
        # Allow development bypass via X-DEV-KEY header or DEV_SEARCH_KEY env (only for dev/testing)
        dev_key = request.headers.get('X-DEV-KEY') or request.args.get('dev_key')
        allowed_dev = app.config.get('DEV_SEARCH_KEY') or os.environ.get('DEV_SEARCH_KEY') or 'devtest'
        if allowed_dev and dev_key and dev_key == allowed_dev:
            return func(*args, **kwargs)
        # app.logger.debug(f'is_connected: session = {session}')
        if 'login_id' not in session:
            error = 'Restricted access! Please authenticate.'
            # Si requête AJAX / JSON, retourner 401 JSON utile au JS
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest' or request.is_json:
                from flask import jsonify
                return jsonify({ 'error': error }), 401
            return render_template('login.html', error=error)
            # return redirect(url_for('login.html'))  # Remplacez 'login.html' par l'URL de votre page de connexion
        return func(*args, **kwargs)

    return wrapper


def _get_github_rate_info(token: str) -> dict | None:
    """Interroge l'endpoint /rate_limit de GitHub et retourne un dict simple
    {'remaining': int, 'limit': int, 'reset': int} ou None en cas d'erreur.
    Utilisé uniquement pour afficher des messages utilisateur plus précis.
    """
    if not token:
        return None
    try:
        import requests
        headers = {"Authorization": f"Bearer {token}", "Accept": "application/json"}
        r = requests.get('https://api.github.com/rate_limit', headers=headers, timeout=2)
        if r.status_code != 200:
            return None
        data = r.json()
        core = data.get('resources', {}).get('core') or data.get('rate') or {}
        return {
            'remaining': int(core.get('remaining', 0)),
            'limit': int(core.get('limit', 0)),
            'reset': int(core.get('reset', 0)),
        }
    except Exception:
        return None


def _format_github_rate_warning(token: str, default_missing_msg: str) -> str:
    """Retourne un message d'avertissement adapté selon la présence du token et du rate info."""
    if not token:
        return default_missing_msg
    info = _get_github_rate_info(token)
    if not info:
        return 'GITHUB_TOKEN configuré mais impossible de vérifier le quota — aperçu généré sans IA.'
    # Calculer temps restant jusqu'au reset
    try:
        reset_ts = int(info.get('reset', 0))
        reset_dt = datetime.utcfromtimestamp(reset_ts)
        minutes = max(0, int((reset_dt - datetime.utcnow()).total_seconds() / 60))
    except Exception:
        minutes = None
    if minutes is None:
        return f"Quota restant: {info.get('remaining')}/{info.get('limit')} — aperçu généré sans IA."
    return f"Quota GitHub remaining: {info.get('remaining')}/{info.get('limit')} (reset dans ~{minutes} min) — aperçu généré sans IA."

def is_admin(func):
    @wraps(func)
    def wrapper(*args, **kwargs):
        user = get_user_by_id(session['login_id'])
        if not user.is_admin:
            error = 'Insufficient privileges for this operation! Please contact administrator...'
            # Si requête AJAX / JSON, retourner 403 JSON utile au JS
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest' or request.is_json:
                from flask import jsonify
                return jsonify({ 'error': error }), 403
            return render_template('login.html', error=error)
        return func(*args, **kwargs)
    return wrapper

def is_editor_or_admin(func):
    @wraps(func)
    def wrapper(*args, **kwargs):
        user = get_user_by_id(session['login_id'])
        # Nouveau : rôle 'Utilisateur' remplace l'ancien 'Editeur'. Autorise donc
        # les appels si l'utilisateur est admin ou 'user'.
        if not (user.is_user or user.is_admin):
            error = 'Insufficient privileges for this operation! Please contact administrator...'
            return render_template('login.html', error=error)
        return func(*args, **kwargs)
    return wrapper

def get_client_ip():
    # Check headers in order of reliability
    if request.headers.getlist("X-Forwarded-For"):
        client_ip = request.headers.getlist("X-Forwarded-For")[0]
    elif request.headers.get("X-Real-IP"):
        client_ip = request.headers.get("X-Real-IP")
    elif request.headers.get("CF-Connecting-IP"):    # Cloudflare
        client_ip = request.headers.get("CF-Connecting-IP")
    else:
        client_ip = request.remote_addr
    return client_ip


def _resolve_user_cv_data(user_id: int) -> dict:
    """Charge le CV JSON par défaut de l'utilisateur si disponible.
    Retourne un dict (chargé) ou None si absent.
    """
    try:
        cvs_dir = Path(app.static_folder) / 'uploads' / 'users' / str(user_id) / 'cvs'
        meta_path = cvs_dir / 'cvs_meta.json'
        meta = {}
        if meta_path.exists():
            import json as _json
            meta = _json.loads(meta_path.read_text(encoding='utf-8'))
        default = meta.get('_default')
        if not default:
            return None
        target = cvs_dir / default
        if target.exists() and target.suffix.lower() == '.json':
            return _json.loads(target.read_text(encoding='utf-8'))
        # If pdf, try to load parsed sibling
        candidate2 = cvs_dir / (target.stem + '.json')
        candidate1 = cvs_dir / 'cv_linkedin_parsed.json'
        for c in (candidate2, candidate1):
            if c.exists():
                try:
                    return _json.loads(c.read_text(encoding='utf-8'))
                except Exception:
                    continue
    except Exception:
        return None
    return None

@app.route('/get_ip')
def get_ip():
    # Obtenir l'adresse IP du client
    # client_ip = request.remote_addr
    # app.logger.debug(f'old client_ip = {client_ip}')
    client_ip = get_client_ip()
    app.logger.debug(f'new client_ip = {client_ip}')

    # Obtenir le nom d'hôte du serveur
    server_hostname = request.host.split(':')[0]

    # Résoudre le nom d'hôte en adresse IP
    server_ip = gethostbyname(server_hostname)

    return f"Adresse IP du client : {client_ip}\nAdresse IP du serveur : {server_ip}"


@app.route('/user_cv_default', methods=['GET'])
@is_connected
def user_cv_default():
    """Retourne le nom du fichier CV par défaut pour l'utilisateur connecté.
    JSON: { default: 'cv.pdf', meta: { ... } }
    """
    user = get_user_by_id(session['login_id'])
    cvs_dir = Path(app.static_folder) / 'uploads' / 'users' / str(user.id) / 'cvs'
    meta_path = cvs_dir / 'cvs_meta.json'
    import json as _json
    meta = {}
    try:
        if meta_path.exists():
            meta = _json.loads(meta_path.read_text(encoding='utf-8'))
    except Exception:
        meta = {}
    default = meta.get('_default')
    # Build full URL when possible to help the client display a direct link
    default_url = None
    try:
        if default:
            rel = f'uploads/users/{user.id}/cvs/{default}'
            default_url = url_for('static', filename=rel)
    except Exception:
        default_url = None
    return jsonify({'default': default, 'default_url': default_url, 'meta': meta})


@app.route('/__debug_user_cv_default/<int:user_id>', methods=['GET'])
def debug_user_cv_default(user_id):
    """Debug endpoint (no auth) returning the same payload as `/user_cv_default` for a given user id.
    Temporary: used to verify client behaviour for tests/dev. Remove before production use.
    """
    cvs_dir = Path(app.static_folder) / 'uploads' / 'users' / str(user_id) / 'cvs'
    meta_path = cvs_dir / 'cvs_meta.json'
    import json as _json
    meta = {}
    try:
        if meta_path.exists():
            meta = _json.loads(meta_path.read_text(encoding='utf-8'))
    except Exception:
        meta = {}
    default = meta.get('_default')
    default_url = None
    try:
        if default:
            rel = f'uploads/users/{user_id}/cvs/{default}'
            default_url = url_for('static', filename=rel)
    except Exception:
        default_url = None
    return jsonify({'default': default, 'default_url': default_url, 'meta': meta})


@app.route('/__debug_preview_cv_data/<int:job_id>', methods=['POST'])
def debug_preview_cv_data(job_id):
    """Debug preview endpoint without auth. Accepts JSON body { "user_id": X, "default_cv": "..." }.
    Returns a light preview (basics.label, summary) and cv_source info for verification.
    """
    payload = request.get_json(silent=True) or {}
    user_id = payload.get('user_id')
    default_cv = payload.get('default_cv')
    # try to load meta if default not provided
    if not default_cv and user_id:
        cvs_dir = Path(app.static_folder) / 'uploads' / 'users' / str(user_id) / 'cvs'
        meta_path = cvs_dir / 'cvs_meta.json'
        try:
            if meta_path.exists():
                meta_obj = _json.loads(meta_path.read_text(encoding='utf-8'))
                default_cv = meta_obj.get('_default')
        except Exception:
            default_cv = None

    # resolve cv_data
    cv_data = None
    if default_cv and user_id:
        try:
            user_cvs = Path(app.static_folder) / 'uploads' / 'users' / str(user_id) / 'cvs'
            target = user_cvs / default_cv
            if target.suffix.lower() == '.json' and target.exists():
                cv_data = _json.loads(target.read_text(encoding='utf-8'))
            elif target.suffix.lower() == '.pdf':
                candidate2 = user_cvs / (target.stem + '.json')
                candidate1 = user_cvs / 'cv_linkedin_parsed.json'
                for c in (candidate2, candidate1):
                    if c.exists():
                        try:
                            cv_data = _json.loads(c.read_text(encoding='utf-8'))
                            break
                        except Exception:
                            continue
        except Exception:
            cv_data = None

    if not cv_data:
        cv_data = load_cv_data(app.static_folder)

    resolved_name = default_cv if default_cv and cv_data else None
    try:
        resolved_url = url_for('static', filename=f'uploads/users/{user_id}/cvs/{resolved_name}') if resolved_name else None
    except Exception:
        resolved_url = None

    basics = cv_data.get('basics', {}) if isinstance(cv_data, dict) else {}
    resp = {
        'cv_title': basics.get('label'),
        'summary': basics.get('summary'),
        'cv_source': resolved_name,
        'cv_source_url': resolved_url,
    }
    return jsonify(resp)


@login_manager.user_loader
def load_user(user_id):
    return User.query.get(int(user_id))


@app.route('/')
def welcome():
    user = None
    token = None
    if 'login_id' in session:
        user = get_user_by_id(session['login_id'])
        # Générer un jeton de récupération de mot de passe
        s = Serializer(app.config['SECRET_KEY'])
        token = s.dumps({'user_id': user.id})
        # Mettez à jour le modèle d'utilisateur avec le jeton et le délai d'expiration
        # user.recovery_token = generate_password_hash(token, method='sha256')
        user.recovery_token = generate_password_hash(token)
        user.token_expiration = datetime.now() + timedelta(hours=24)
        db.session.commit()
    else:
        # client_ip = request.remote_addr
        client_ip = get_client_ip()
        user_agent_string = request.headers.get('User-Agent')
        user_agent: UserAgent = parse(user_agent_string)
        browser_info = f"Family = {user_agent.browser.family}, Version = {user_agent.browser.version_string}"
        app.logger.debug(f"Client IP: {client_ip}, Browser: ({browser_info})")
    return render_template('index.html', session=session, user=user, token=token)


@app.route("/register", methods=['GET', 'POST'])
def register():
    # Logique d'enregistrement ici
    error = None
    if request.method == 'POST':
        if request.form['confirm_password'] != request.form['password']:
            # flash('Incorrect login credentials.', 'error')
            error = 'Password does not match! Please try again.'
        else:
            username: str = request.form['username']
            email: str = request.form['email']
            existing_user: User = User.query.filter_by(username=username).first()
            existing_email: User = User.query.filter_by(email=email).first()
            if existing_user:
                error = f'user {existing_user.username} already exists! Please choose another name.'
            elif existing_email:
                error = f'email {existing_email.email} already exists! Please choose another email.'
            elif not check(app.config['REGEX'], email):
                error = f'email {email} is invalid! Please check syntax.'
            else:
                user = User(username=username, password=request.form['password'],
                            creation_date=datetime.now(), email=email)
                # logging.warning("See this message in Flask Debug Toolbar!")
                db.session.add(user)
                db.session.commit()
                s = Serializer(app.config['SECRET_KEY'])
                # s = URLSafeSerializer('SECRET_KEY')
                token = s.dumps({'user_id': user.id})

                # Mettez à jour le modèle d'utilisateur avec le jeton et le délai d'expiration
                # user.recovery_token = generate_password_hash(token, method='sha256')
                user.recovery_token = generate_password_hash(token)
                user.token_expiration = datetime.now() + timedelta(minutes=10)
                # app.logger.debug(f'time zone info: {user.token_expiration.tzinfo}')

                db.session.commit()

                # Envoyer un e-mail de confirmation d'inscription

                flash('Un e-mail de demande de confirmation d\'inscription a été envoyé!', 'success')
                confirm_link = url_for('validate_email', token=token, _external=True)
                send_confirmation_email(app=app, confirm_link=confirm_link, user=user,
                                        author=app.config['GMAIL_FULLNAME'], cv_resume=app.config['CV_RESUME'])
                return redirect(url_for('register'))

    return render_template('register.html', error=error)


@app.route("/login", methods=['GET', 'POST'])
def login():
    # Logique de connexion ici
    error = None
    if request.method == 'POST':
        user = User.query.filter_by(username=request.form['username']).first()
        app.logger.debug(f'user = {user} - password: {user.password} - clear pwd = {request.form["password"]}')
        if user and check_password_and_upgrade(user, password=request.form['password']):
            if user.validated:
                session['login_id'] = user.id
                app.logger.debug(f'user (login) = {user.username} - id = {user.id} - session: {session}')
                # client_ip = request.remote_addr
                client_ip = get_client_ip()
                user_agent_string = request.headers.get('User-Agent')
                user_agent: UserAgent = parse(user_agent_string)
                sess = Session(login_id=user.id, start=datetime.now(), client_ip=client_ip,
                               browser_family=user_agent.browser.family,
                               browser_version=user_agent.browser.version_string)
                db.session.add(sess)
                db.session.commit()
                return redirect(url_for('welcome'))
            else:
                error = 'Your account is not yet validated! Please check your email for the confirmation link.'
                return render_template('login.html', error=error)
        else:
            error = 'Incorrect login credentials. Please try again.'
            return render_template('login.html', error=error)
    return render_template('login.html', error=error)


@app.route("/change_password", methods=['GET', 'POST'])
@is_connected
def change_password():
    error = None
    if request.method == 'POST':
        user = get_user_by_id(session['login_id'])
        new_password: str = request.form["new_password"]
        confirm_new_password: str = request.form["confirm_new_password"]
        app.logger.debug(
            f'user (change pwd) = {user.username} - new pwd = {new_password} - confirm new_pwd = {confirm_new_password}')
        if new_password == confirm_new_password:
            # user.password = generate_password_hash(new_password, method='sha256')
            user.password = generate_password_hash(new_password)
            db.session.add(user)
            db.session.commit()
            flash('Password was successfully changed!')
            return redirect(url_for('welcome'))
        else:
            error = 'Passwords does not match! Please try again.'
    return render_template('reset_password.html', error=error)


@app.route('/request_reset_password', methods=['GET', 'POST'])
def request_reset_password():
    if request.method == 'POST':
        email = request.form.get('email')
        user = User.query.filter_by(email=email.lower()).first()
        if user:
            # Générer un jeton de récupération de mot de passe
            s = Serializer(app.config['SECRET_KEY'])
            token = s.dumps({'user_id': user.id})

            # Mettez à jour le modèle d'utilisateur avec le jeton et le délai d'expiration
            # user.recovery_token = generate_password_hash(token, method='sha256')
            user.recovery_token = generate_password_hash(token)
            user.token_expiration = datetime.now() + timedelta(minutes=10)

            db.session.commit()

            # Envoyer le lien de récupération par e-mail (vous devez implémenter cette partie)
            # Vous pouvez utiliser un package comme Flask-Mail pour envoyer des e-mails.

            flash('Un e-mail de récupération de mot de passe a été envoyé.', 'success')
            reset_link = url_for('reset_password', token=token, _external=True)
            send_password_recovery_email(app=app, reset_link=reset_link, user=user, author=app.config['GMAIL_FULLNAME'],
                                         cv_resume=app.config['CV_RESUME'])
            return redirect(url_for('login'))

        flash('Aucun utilisateur trouvé avec cet e-mail.', 'error')

    return render_template('request_reset_password.html')


@app.route('/validate_email/<token>', methods=['GET', 'POST'])
def validate_email(token):
    error: str = None
    # Vérifier si le jeton est valide
    s = Serializer(app.config['SECRET_KEY'])
    try:
        data = s.loads(token)
        user = User.query.get_or_404(data['user_id'])
        # Calculate the time difference
        remaining_minutes = int((user.token_expiration - datetime.now()).total_seconds() / 60)
        app.logger.debug(f'remaining minutes: {remaining_minutes}')
        if remaining_minutes <= 0:
            raise Exception
    except Exception as e:
        app.logger.debug(e)
        flash('Le lien de confirmation d\'inscription est invalide ou a expiré.')
        user = User.query.get_or_404(data.get('user_id'))
        if user:
            db.session.delete(user)
            db.session.commit()
        return redirect(url_for('login'))

    user = User.query.get(data['user_id'])

    # Mettre à jour le champ de confirmation d'inscription de l'utilisateur
    user.validated = True

    # Réinitialiser le champ de récupération de mot de passe
    user.recovery_token = None
    user.token_expiration = None

    db.session.commit()

    flash('Votre compte a été confirmé avec succès.', 'success')
    return redirect(url_for('login'))


@app.route('/reset_password/<token>', methods=['GET', 'POST'])
def reset_password(token):
    error: str = None
    # Vérifier si le jeton est valide
    s = Serializer(app.config['SECRET_KEY'])
    try:
        data = s.loads(token)
        user = User.query.get_or_404(data['user_id'])
        # Calculate the time difference
        remaining_minutes = int((user.token_expiration - datetime.now()).total_seconds() / 60)
        app.logger.debug(f'remaining minutes: {remaining_minutes}')
        if remaining_minutes <= 0:
            raise Exception
    except:
        flash('Le lien de réinitialisation de mot de passe est invalide ou a expiré.')
        return redirect(url_for('login'))

    user = User.query.get(data['user_id'])
    app.logger.debug(f'reset password user {user.username} - data = {data} \n - token = {token}')

    if request.method == 'POST':

        new_password = request.form.get('new_password')
        confirm_new_password = request.form.get('confirm_new_password')

        if new_password == confirm_new_password:
            # Mettre à jour le mot de passe de l'utilisateur
            # user.password = generate_password_hash(new_password, method='sha256')
            user.password = generate_password_hash(new_password)

            # Réinitialiser le champ de récupération de mot de passe
            user.recovery_token = None
            user.token_expiration = None

            db.session.commit()

            flash('Le mot de passe a été réinitialisé avec succès.', 'success')
            return redirect(url_for('login'))
        else:
            error = 'Les mots de passe ne correspondent pas.'

    return render_template('reset_password.html', error=error, token=token)


@app.route("/logout")
@is_connected
def logout():
    user = get_user_by_id(session['login_id'])
    sess = get_session_by_login(username=user.username)
    if sess is not None:
        sess.end = datetime.now()
        db.session.commit()
    logout_user()
    # clear all session data including cookies
    session.clear()
    flash('You have been logged out.', 'success')
    return redirect(url_for('login'))


@app.route('/accounts')
@is_connected
@is_admin
def show_accounts():
    user = get_user_by_id(session['login_id'])
    # Query users together with their session counts and sort by count desc
    count_expr = func.count(Session.id)
    results = db.session.query(User, count_expr.label('connection_count')).outerjoin(Session, Session.login_id == User.id).group_by(User.id).order_by(desc(count_expr)).all()
    accounts = []
    for user_obj, conn_count in results:
        # attach the count and last session
        user_obj.connection_count = conn_count
        user_obj.last_session = Session.query.filter_by(login_id=user_obj.id).order_by(desc(Session.start)).first()
        accounts.append(user_obj)
    return render_template('accounts.html', accounts=accounts, user=user)


@app.route('/delete_session/<int:id>', methods=['POST'])
@is_connected
@is_admin
def delete_session(id):
    session_obj = Session.query.get_or_404(id)
    current_user_id = session['login_id']
    if session_obj.end is None:
        if session_obj.login_id == current_user_id:
            return jsonify({ 'error': 'Cannot close your own active session' }), 403
        # Close active session of another user
        session_obj.end = datetime.now()
        db.session.commit()
    else:
        # Delete closed session permanently (any user)
        db.session.delete(session_obj)
        db.session.commit()
    return '', 200


@app.route('/delete_sessions', methods=['GET', 'POST'])
@is_connected
@is_admin
def delete_sessions():
    """Supprime (marque end=now) plusieurs sessions à la fois.
    - GET : renvoie un message de diagnostic (utilisé pour vérifier que la route est bien active).
    - POST : attend un JSON { "ids": [1,2,3] } et met à jour les sessions.
    """
    app.logger.debug(f"delete_sessions called: method={request.method} path={request.path}")
    try:
        raw = request.get_data(as_text=True)
    except Exception:
        raw = None
    app.logger.debug(f"delete_sessions raw payload: {raw}")

    if request.method == 'GET':
        return jsonify({ 'message': 'Endpoint /delete_sessions disponible', 'method': 'GET' }), 200

    # POST handler
    try:
        data = request.get_json(force=True)
    except Exception:
        return jsonify({ 'error': 'Invalid JSON' }), 400

    if not data or 'ids' not in data:
        return jsonify({ 'error': 'Missing ids list' }), 400

    ids = data.get('ids')
    if not isinstance(ids, list) or any(not isinstance(i, int) for i in ids):
        return jsonify({ 'error': 'ids must be a list of integers' }), 400

    # Query all sessions matching provided ids
    sessions_to_close = Session.query.filter(Session.id.in_(ids)).all()
    print(f"DEBUG delete_sessions: found {len(sessions_to_close)} sessions for ids {ids}")
    now = datetime.now()
    closed_ids = []
    deleted_ids = []
    current_user_id = session['login_id']
    for sess in sessions_to_close:
        print(f"DEBUG Processing session {sess.id}, end={sess.end}, login_id={sess.login_id}")
        if sess.end is None:
            if sess.login_id == current_user_id:
                print(f"DEBUG Skipping own active session {sess.id}")
                continue  # Skip closing own session
            sess.end = now
            closed_ids.append(sess.id)
        else:
            db.session.delete(sess)
            deleted_ids.append(sess.id)
    try:
        db.session.commit()
        print(f"DEBUG Committed successfully: closed {closed_ids}, deleted {deleted_ids}")
    except Exception as e:
        print(f"DEBUG Commit failed: {e}")
        db.session.rollback()
        return jsonify({ 'error': 'Commit failed' }), 500
    return jsonify({ 'closed': closed_ids, 'deleted': deleted_ids }), 200


@app.route('/sessions')
@is_connected
@is_admin
def show_sessions():
    # Reverse order query
    sessions = Session.query.filter(Session.end.is_(None)).order_by(desc(Session.id)).all()
    user = get_user_by_id(session['login_id'])
    return render_template('sessions.html', sessions=sessions, user=user)


@app.route('/closed_sessions')
@is_connected
@is_admin
def show_closed_sessions():
    username_filter = request.args.get('username', '')
    query = Session.query.filter(Session.end.is_not(None))
    if username_filter:
        query = query.join(User).filter(User.username == username_filter)
    sessions = query.order_by(desc(Session.end)).all()
    users = db.session.query(User.username).distinct().order_by(User.username).all()
    return render_template('closed_sessions.html', sessions=sessions, username_filter=username_filter, users=users)


@app.route('/suivi')
@is_connected
def show_all():
    user_id = session['login_id']
    app.logger.debug('This is a debug message.')
    # Allow selecting to view another user's offers via query param 'view_user'
    view_user = request.args.get('view_user')
    from sqlalchemy import or_
    if view_user:
        try:
            target_id = int(view_user)
        except Exception:
            target_id = None
        # authorize viewing: either target is current user or target allowed sharing
        if target_id and (target_id == user_id or (User.query.get(target_id) and User.query.get(target_id).allow_view_offers)):
            jobs = Job.query.filter(Job.user_id == target_id).all()
        else:
            flash('Action non autorisée : cet utilisateur ne partage pas ses offres.', 'error')
            return redirect(url_for('show_all'))
    else:
        # default: show only the connected user's jobs
        jobs = Job.query.filter(Job.user_id == user_id).all()
    def most_recent_date(job):
        dates = [d for d in (job.relaunchDate, job.applicationDate) if d]
        if not dates:
            return datetime.min
        return max(dates)

    jobs.sort(key=most_recent_date, reverse=True)
    for j in jobs:
        rd = most_recent_date(j)
        try:
            j.recent_ts = int(rd.timestamp())
        except Exception:
            j.recent_ts = 0
    user = get_user_by_id(user_id)
    # Determine which jobs already have a saved CV PDF on disk
    uploads_dir = Path(app.static_folder) / 'uploads'
    cv_pdf_urls = {}
    # check per-user generated dir first
    if uploads_dir.exists():
        # old global files cv_{id}.pdf
        for p in uploads_dir.glob('cv_*.pdf'):
            stem = p.stem.replace('cv_', '')
            if stem.isdigit():
                try:
                    jid = int(stem)
                    rel = f'uploads/{p.name}'
                    cv_pdf_urls[jid] = url_for('static', filename=rel)
                except Exception:
                    continue
    # per-user generated files
    users_dir = uploads_dir / 'users'
    if users_dir.exists():
        for user_dir in users_dir.iterdir():
            gen_dir = user_dir / 'generated'
            if not gen_dir.exists():
                continue
            for p in gen_dir.glob('cv_*.pdf'):
                stem = p.stem.replace('cv_', '')
                if stem.isdigit():
                    try:
                        jid = int(stem)
                        rel = f'uploads/users/{user_dir.name}/generated/{p.name}'
                        cv_pdf_urls[jid] = url_for('static', filename=rel)
                    except Exception:
                        continue
    # Charger formations et certifications pour les modales avancées
    try:
        # Prefer user's default CV JSON when available (so modales avancées utilisent le bon profil)
        _cvd = _resolve_user_cv_data(user.id) or load_cv_data(app.static_folder)
        cv_education = [
            {
                'idx': i,
                'label': ' - '.join(p for p in [e.get('studyType', ''), e.get('area', '')] if p)
                         or e.get('institution', ''),
                'institution': e.get('institution', ''),
                'year': (e.get('endDate') or '')[:4],
            }
            for i, e in enumerate(_cvd.get('education', []))
        ]
        cv_certificates = [
            {
                'name': c['name'],
                'issuer': c.get('issuer', ''),
                'year': (c.get('date') or '')[:4],
            }
            for c in _cvd.get('certificates', [])
        ]
        cv_skills = [
            {
                'name': s['name'],
                'level': s.get('level', ''),
                'rating': s.get('rating', 0),
            }
            for s in _cvd.get('skills', [])
        ]
        cv_references = [
            {
                'name': r['name'],
                'excerpt': (r.get('reference') or '')[:120].rstrip(),
            }
            for r in _cvd.get('references', [])
        ]
        cv_projects = [
            {
                'name': p['name'],
                'lang': p.get('primaryLanguage', ''),
                'desc': (p.get('summary') or p.get('description') or '')[:80],
            }
            for p in _cvd.get('projects', [])
        ]
    except Exception:
        cv_education = []
        cv_certificates = []
        cv_skills = []
        cv_references = []
        cv_projects = []
    # Prepare list of users for the selector: include current user and users who opted-in
    users_list = User.query.filter((User.id == user_id) | (User.allow_view_offers == True)).order_by(User.username).all()
    return render_template('candidatures.html', jobs=jobs, user=user, users=users_list, cv_pdf_urls=cv_pdf_urls,
                           cv_education=cv_education, cv_certificates=cv_certificates,
                           cv_skills=cv_skills, cv_references=cv_references, cv_projects=cv_projects)


@app.route('/generate_cover_letter_pdf/<int:id>', endpoint='generate_cover_letter_pdf')
@is_connected
def generate_cover_letter_pdf(id):
    # ...existing code...
    job = Job.query.get_or_404(id)
    try:
        template_path = resolve_cover_letter_template_path(app.static_folder)
        letter_date = datetime.now(app.config['PARIS']) if app.config.get('PARIS') else datetime.now()
        pdf_bytes = generate_cover_letter_pdf_bytes(job=job, template_path=template_path, letter_date=letter_date)
    except FileNotFoundError:
        flash('Le template Word de lettre de motivation est introuvable.', 'error')
        return redirect(url_for('show_all'))
    except RuntimeError as exc:
        app.logger.error(f'Erreur lors de la génération du PDF pour la candidature #{job.id}: {exc}')
        flash(str(exc), 'error')
        return redirect(url_for('show_all'))

    return send_file(
        BytesIO(pdf_bytes),
        mimetype='application/pdf',
        as_attachment=True,
        download_name=build_cover_letter_pdf_filename(job),
    )


@app.route('/preview_cv_data/<int:id>', methods=['POST'])
@is_connected
def preview_cv_data(id):
    """Retourne les suggestions IA de personnalisation du CV au format JSON (pour l'aperçu modal)."""
    job = Job.query.get_or_404(id)
    github_token = app.config.get('GITHUB_TOKEN', '')
    payload = request.get_json(force=True) or {}
    additional_prompt = payload.get('additional_prompt', '')
    ats_mode = bool(payload.get('ats_mode', False))
    if ats_mode:
        additional_prompt = _ft_ats_prefix(True) + additional_prompt
    include_sections = payload.get('include_sections') or []
    # Sélections individuelles formations / certifications
    selected_education = payload.get('selected_education')
    selected_certificates = payload.get('selected_certificates')
    selected_skills = payload.get('selected_skills')
    selected_references = payload.get('selected_references')
    selected_projects = payload.get('selected_projects')

    def _fallback(cv_data, warning=''):
        sugg = {
            'cv_title': cv_data.get('basics', {}).get('label', 'Développeur / Consultant IT'),
            'summary': cv_data.get('basics', {}).get('summary', ''),
            'highlighted_work_indices': list(range(min(4, len(cv_data.get('work', []))))),
            'highlighted_skill_names': [s['name'] for s in cv_data.get('skills', [])[:6]],
            'warning': warning or _format_github_rate_warning(github_token, 'GITHUB_TOKEN non configuré : aperçu généré sans IA.'),
            'source': 'fallback',
        }
        wl = cv_data.get('work', [])
        sugg['highlighted_work_details'] = [
            {
                'position': wl[i].get('position', ''),
                'company':  wl[i].get('name', ''),
                'dates':    f"{(wl[i].get('startDate') or '')[:7]} – {(wl[i].get('endDate') or '')[:7] or 'présent'}",
            }
            for i in sugg['highlighted_work_indices'] if 0 <= i < len(wl)
        ]
        return sugg

    def _resolve_default_cv_data(user_id: int, default_name: str) -> dict | None:
        """Tente de charger un CV JSON correspondant au nom fourni dans les uploads utilisateur.
        Retourne un dict si trouvé, sinon None.
        """
        try:
            user_cvs = Path(app.static_folder) / 'uploads' / 'users' / str(user_id) / 'cvs'
            if not user_cvs.exists():
                return None
            target = user_cvs / default_name
            # If it's a JSON file, load directly
            if target.suffix.lower() == '.json' and target.exists():
                return _json.loads(target.read_text(encoding='utf-8'))
            # If it's a PDF, try to find a parsed JSON sibling
            if target.suffix.lower() == '.pdf':
                # common parsed filename used by import_cv
                candidate1 = user_cvs / 'cv_linkedin_parsed.json'
                candidate2 = user_cvs / (target.stem + '.json')
                for c in (candidate2, candidate1):
                    if c.exists():
                        try:
                            return _json.loads(c.read_text(encoding='utf-8'))
                        except Exception:
                            continue
            return None
        except Exception:
            return None

    try:
        # Determine which CV to use: prefer explicit payload.default_cv, else use user's meta _default if present
        default_cv = payload.get('default_cv')
        if not default_cv:
            # try to read user's cvs_meta.json
            try:
                meta_path = Path(app.static_folder) / 'uploads' / 'users' / str(job.user_id) / 'cvs' / 'cvs_meta.json'
                if meta_path.exists():
                    meta_obj = _json.loads(meta_path.read_text(encoding='utf-8'))
                    default_cv = meta_obj.get('_default')
            except Exception:
                default_cv = None

        if default_cv:
            resolved = _resolve_default_cv_data(job.user_id, default_cv)
            if resolved:
                cv_data = resolved
            else:
                cv_data = load_cv_data(app.static_folder)
        else:
            cv_data = load_cv_data(app.static_folder)
        # expose resolved name for debugging / client
        # expose which filename was used (for debugging/client display)
        resolved_name = default_cv if (default_cv and resolved) else (default_cv if default_cv and not resolved else None)
        app.logger.info(f"preview_cv_data: resolved default_cv='{resolved_name}' for user_id={job.user_id} job_id={id}")
        if github_token:
            suggestions = get_ai_cv_suggestions(
                job=job, cv_data=cv_data, github_token=github_token,
                additional_prompt=additional_prompt, include_sections=include_sections,
                selected_education=selected_education, selected_certificates=selected_certificates,
                selected_skills=selected_skills, selected_references=selected_references,
                selected_projects=selected_projects,
            )
        else:
            suggestions = _fallback(cv_data)
            suggestions['_active_sections'] = include_sections
        work_list = cv_data.get('work', [])
        hl_indices = suggestions.get('highlighted_work_indices') or []
        suggestions['highlighted_work_details'] = [
            {
                'position': work_list[i].get('position', ''),
                'company':  work_list[i].get('name', ''),
                'dates':    f"{(work_list[i].get('startDate') or '')[:7]} – {(work_list[i].get('endDate') or '')[:7] or 'présent'}",
            }
            for i in hl_indices if 0 <= i < len(work_list)
        ]
        # Labels lisibles des sections actives pour affichage dans la modale
        _section_labels = {
            'education': '🎓 Formations', 'certificates': '🏅 Certifications',
            'skills_rating': '⭐ Niveaux compétences', 'references': '💬 Références',
            'github_projects': '🐙 Projets GitHub',
        }
        suggestions['active_section_labels'] = [
            _section_labels[s] for s in (suggestions.get('_active_sections') or []) if s in _section_labels
        ]
        # include cv_source so client can show which CV was used for the preview
        resp = suggestions.copy() if isinstance(suggestions, dict) else {'suggestions': suggestions}
        resp['cv_source'] = resolved_name
        try:
            if resolved_name:
                rel = f'uploads/users/{job.user_id}/cvs/{resolved_name}'
                resp['cv_source_url'] = url_for('static', filename=rel)
            else:
                resp['cv_source_url'] = None
        except Exception:
            resp['cv_source_url'] = None
        return jsonify(resp)
    except FileNotFoundError as exc:
        return jsonify({'error': str(exc)}), 404
    except Exception as exc:
        app.logger.error(f'preview_cv_data error: {exc}')
        try:
            cv_data = load_cv_data(app.static_folder)
            return jsonify(_fallback(cv_data, warning=f"Aperçu généré sans IA : {exc}")), 200
        except Exception:
            return jsonify({'error': f"Erreur lors de l'appel IA : {exc}"}), 500


@app.route('/save_cv_pdf/<int:id>', methods=['POST'], endpoint='save_cv_pdf')
@is_connected
def save_cv_pdf(id):
    """Génère le CV personnalisé IA, le sauvegarde sur disque et retourne l'URL de téléchargement."""
    job = Job.query.get_or_404(id)
    github_token = app.config.get('GITHUB_TOKEN', '')
    payload = request.get_json(force=True) or {}
    additional_prompt = payload.get('additional_prompt', '')
    ats_mode = bool(payload.get('ats_mode', False))
    if ats_mode:
        additional_prompt = _ft_ats_prefix(True) + additional_prompt
    include_sections = payload.get('include_sections') or []
    selected_education = payload.get('selected_education')
    selected_certificates = payload.get('selected_certificates')
    selected_skills = payload.get('selected_skills')
    selected_references = payload.get('selected_references')
    selected_projects = payload.get('selected_projects')

    def _fallback_suggestions(cv_data, warning=''):
        return {
            'cv_title': cv_data.get('basics', {}).get('label', 'Développeur / Consultant IT'),
            'summary': cv_data.get('basics', {}).get('summary', ''),
            'highlighted_work_indices': list(range(min(4, len(cv_data.get('work', []))))),
            'highlighted_skill_names': [s['name'] for s in cv_data.get('skills', [])[:6]],
            'warning': warning or _format_github_rate_warning(github_token, 'GITHUB_TOKEN non configuré.'),
            'source': 'fallback',
        }

    def _resolve_default_cv_data(user_id: int, default_name: str) -> dict | None:
        try:
            user_cvs = Path(app.static_folder) / 'uploads' / 'users' / str(user_id) / 'cvs'
            if not user_cvs.exists():
                return None
            target = user_cvs / default_name
            if target.suffix.lower() == '.json' and target.exists():
                return _json.loads(target.read_text(encoding='utf-8'))
            if target.suffix.lower() == '.pdf':
                candidate1 = user_cvs / 'cv_linkedin_parsed.json'
                candidate2 = user_cvs / (target.stem + '.json')
                for c in (candidate2, candidate1):
                    if c.exists():
                        try:
                            return _json.loads(c.read_text(encoding='utf-8'))
                        except Exception:
                            continue
            return None
        except Exception:
            return None

    try:
        default_cv = payload.get('default_cv')
        if not default_cv:
            try:
                meta_path = Path(app.static_folder) / 'uploads' / 'users' / str(job.user_id) / 'cvs' / 'cvs_meta.json'
                if meta_path.exists():
                    meta_obj = _json.loads(meta_path.read_text(encoding='utf-8'))
                    default_cv = meta_obj.get('_default')
            except Exception:
                default_cv = None

        if default_cv:
            resolved = _resolve_default_cv_data(job.user_id, default_cv)
            if resolved:
                cv_data = resolved
            else:
                cv_data = load_cv_data(app.static_folder)
        else:
            cv_data = load_cv_data(app.static_folder)
        if github_token:
            suggestions = get_ai_cv_suggestions(
                job=job, cv_data=cv_data, github_token=github_token,
                additional_prompt=additional_prompt, include_sections=include_sections,
                selected_education=selected_education, selected_certificates=selected_certificates,
                selected_skills=selected_skills, selected_references=selected_references,
                selected_projects=selected_projects,
            )
        else:
            suggestions = _fallback_suggestions(cv_data)
        pdf_bytes = generate_tailored_cv_pdf_bytes(
            job=job, cv_data=cv_data, suggestions=suggestions,
            include_sections=include_sections,
            selected_education=selected_education,
            selected_certificates=selected_certificates,
            selected_skills=selected_skills,
            selected_references=selected_references,
            selected_projects=selected_projects,
        )
    except FileNotFoundError as exc:
        return jsonify({'error': str(exc)}), 404
    except Exception as exc:
        # Log full traceback to help debugging unexpected types inside job/cv fields
        app.logger.exception('save_cv_pdf error')
        try:
            cv_data = load_cv_data(app.static_folder)
            suggestions = _fallback_suggestions(cv_data, warning=str(exc))
            pdf_bytes = generate_tailored_cv_pdf_bytes(
                job=job, cv_data=cv_data, suggestions=suggestions,
                include_sections=include_sections,
                selected_education=selected_education,
                selected_certificates=selected_certificates,
                selected_skills=selected_skills,
                selected_references=selected_references,
                selected_projects=selected_projects,
            )
        except Exception as exc2:
            return jsonify({'error': f'Génération PDF échouée : {exc2}'}), 500

    # Sauvegarder sur disque dans static/uploads/users/{user_id}/generated/
    owner_id = job.user_id
    user_dir = Path(app.static_folder) / 'uploads' / 'users' / str(owner_id) / 'generated'
    user_dir.mkdir(parents=True, exist_ok=True)
    cv_filename = f'cv_{id}.pdf'
    cv_path = user_dir / cv_filename
    cv_path.write_bytes(pdf_bytes)

    # URL relative pour téléchargement
    rel_path = f'uploads/users/{owner_id}/generated/{cv_filename}'
    download_url = url_for('static', filename=rel_path)
    try:
        pdf_name = build_cv_pdf_filename(job)
    except Exception as e:
        app.logger.exception(f'build_cv_pdf_filename failed: {e}')
        # Fallback to a safe filename
        pdf_name = f'cv_{id}.pdf'

    # expose which filename was used (for debugging/client display)
    resolved_name = default_cv if (default_cv and resolved) else (default_cv if default_cv and not resolved else None)
    app.logger.info(f"save_cv_pdf: resolved default_cv='{resolved_name}' for user_id={job.user_id} job_id={id}")
    try:
        if resolved_name:
            rel = f'uploads/users/{job.user_id}/cvs/{resolved_name}'
            resolved_url = url_for('static', filename=rel)
        else:
            resolved_url = None
    except Exception:
        resolved_url = None

    return jsonify({'url': download_url, 'filename': pdf_name, 'cv_source': resolved_name, 'cv_source_url': resolved_url})


@app.route('/generate_cv_pdf/<int:id>', endpoint='generate_cv_pdf')
@is_connected
def generate_cv_pdf(id):
    """Télécharge directement le CV PDF sauvegardé sur disque, ou génère à la volée si absent."""
    job = Job.query.get_or_404(id)
    # First check per-user generated directory
    owner_id = job.user_id
    user_generated = Path(app.static_folder) / 'uploads' / 'users' / str(owner_id) / 'generated' / f'cv_{id}.pdf'
    if user_generated.exists():
        return send_file(
            str(user_generated),
            mimetype='application/pdf',
            as_attachment=True,
            download_name=build_cv_pdf_filename(job),
        )
    # Fallback: older location
    uploads_dir = Path(app.static_folder) / 'uploads'
    cv_path = uploads_dir / f'cv_{id}.pdf'
    if cv_path.exists():
        return send_file(
            str(cv_path),
            mimetype='application/pdf',
            as_attachment=True,
            download_name=build_cv_pdf_filename(job),
        )
    # Fallback : génération à la volée
    github_token = app.config.get('GITHUB_TOKEN', '')
    try:
        # Prefer the connected user's default CV when building ATS keywords so
        # derived keywords reflect the user's profile instead of the global static/cv.json
        cv_data = _resolve_user_cv_data(session.get('login_id')) or load_cv_data(app.static_folder)
        if github_token:
            suggestions = get_ai_cv_suggestions(job=job, cv_data=cv_data, github_token=github_token)
        else:
            suggestions = {
                'cv_title': cv_data.get('basics', {}).get('label', 'Développeur / Consultant IT'),
                'summary': cv_data.get('basics', {}).get('summary', ''),
                'highlighted_work_indices': list(range(min(4, len(cv_data.get('work', []))))),
                'highlighted_skill_names': [s['name'] for s in cv_data.get('skills', [])[:6]],
                'source': 'fallback',
            }
        pdf_bytes = generate_tailored_cv_pdf_bytes(job=job, cv_data=cv_data, suggestions=suggestions)
    except Exception as exc:
        app.logger.error(f'generate_cv_pdf fallback error: {exc}')
        flash(f'Erreur lors de la génération du CV : {exc}', 'error')
        return redirect(url_for('show_all'))

    return send_file(
        BytesIO(pdf_bytes),
        mimetype='application/pdf',
        as_attachment=True,
        download_name=build_cv_pdf_filename(job),
    )


@app.route('/preview_lm_ai/<int:id>', methods=['POST'])
@is_connected
def preview_lm_ai(id):
    """Génère une lettre de motivation par IA et retourne le texte au format JSON."""
    job = Job.query.get_or_404(id)
    github_token = app.config.get('GITHUB_TOKEN', '')
    payload = request.get_json(force=True) or {}
    additional_prompt = payload.get('additional_prompt', '')
    ats_mode = bool(payload.get('ats_mode', False))
    if ats_mode:
        additional_prompt = _ft_ats_prefix(True) + additional_prompt
    include_sections = payload.get('include_sections') or []
    selected_education = payload.get('selected_education')
    selected_certificates = payload.get('selected_certificates')
    selected_skills = payload.get('selected_skills')
    selected_references = payload.get('selected_references')
    selected_projects = payload.get('selected_projects')
    selected_premium_modules = payload.get('selected_premium_modules') or []
    _section_labels = {
        'education': '🎓 Formations', 'certificates': '🏅 Certifications',
        'skills_rating': '⭐ Niveaux compétences', 'references': '💬 Références',
        'github_projects': '🐙 Projets GitHub',
    }
    _premium_labels = {
        'profiles': '🔗 Profils en ligne',
        'languages': '🗣️ Langues',
        'volunteer': '🤝 Bénévolat',
        'awards': '🏆 Distinctions',
        'publications': '📚 Publications',
    }
    try:
        # Allow overriding CV data via default_cv property (client may pass the filename stored for the job owner)
        default_cv = payload.get('default_cv')
        if not default_cv:
            try:
                meta_path = Path(app.static_folder) / 'uploads' / 'users' / str(job.user_id) / 'cvs' / 'cvs_meta.json'
                if meta_path.exists():
                    meta_obj = _json.loads(meta_path.read_text(encoding='utf-8'))
                    default_cv = meta_obj.get('_default')
            except Exception:
                default_cv = None

        def _resolve_default_cv_data(user_id: int, default_name: str) -> dict | None:
            try:
                user_cvs = Path(app.static_folder) / 'uploads' / 'users' / str(user_id) / 'cvs'
                if not user_cvs.exists():
                    return None
                target = user_cvs / default_name
                if target.suffix.lower() == '.json' and target.exists():
                    return _json.loads(target.read_text(encoding='utf-8'))
                if target.suffix.lower() == '.pdf':
                    candidate1 = user_cvs / 'cv_linkedin_parsed.json'
                    candidate2 = user_cvs / (target.stem + '.json')
                    for c in (candidate2, candidate1):
                        if c.exists():
                            try:
                                return _json.loads(c.read_text(encoding='utf-8'))
                            except Exception:
                                continue
                return None
            except Exception:
                return None

        resolved = None
        if default_cv:
            resolved = _resolve_default_cv_data(job.user_id, default_cv)

        # expose resolved name for debugging / client
        resolved_name = default_cv if (default_cv and resolved) else (default_cv if default_cv and not resolved else None)
        app.logger.info(f"preview_lm_ai: resolved default_cv='{resolved_name}' for user_id={job.user_id} job_id={id}")

        text = get_ai_cover_letter_text(
            job=job, github_token=github_token,
            additional_prompt=additional_prompt, include_sections=include_sections,
            selected_education=selected_education,
            selected_certificates=selected_certificates,
            selected_skills=selected_skills,
            selected_references=selected_references,
            selected_projects=selected_projects,
            selected_premium_modules=selected_premium_modules,
            cv_data=resolved,
        )
        resp = {
            'text': text,
            'active_section_labels': [_section_labels[s] for s in include_sections if s in _section_labels],
            'active_premium_labels': [_premium_labels[s] for s in selected_premium_modules if s in _premium_labels],
            'cv_source': resolved_name,
        }
        try:
            if resolved_name:
                rel = f'uploads/users/{job.user_id}/cvs/{resolved_name}'
                resp['cv_source_url'] = url_for('static', filename=rel)
            else:
                resp['cv_source_url'] = None
        except Exception:
            resp['cv_source_url'] = None
        return jsonify(resp)
    except Exception as exc:
        app.logger.error(f'preview_lm_ai error: {exc}')
        status_code = 429 if ('429' in str(exc) or 'Too Many Requests' in str(exc)) else 500
        return jsonify({'error': str(exc)}), status_code


@app.route('/save_lm_text/<int:id>', methods=['POST'])
@is_connected
def save_lm_text(id):
    """Enregistre le texte de LM généré par IA dans la fiche candidature."""
    job = Job.query.get_or_404(id)
    payload = request.get_json(force=True) or {}
    text = payload.get('text', '').strip()
    if not text:
        return jsonify({'error': 'Texte vide'}), 400
    job.cover_letter_text = text
    db.session.commit()
    return jsonify({'ok': True})


@app.route('/new/', methods=['GET', 'POST'])
@is_connected
@is_editor_or_admin
def new():
    if request.method == 'POST':
        email: str = request.form.get('email', '')
        if not (request.form.get('name') and request.form.get('company')):
            flash('Veuillez remplir tous les champs obligatoires (Offre et Entreprise)', 'error')
            user = get_user_by_id(session['login_id'])
            # return the form with previously entered values
            return render_template('new.html', form_data=request.form, user=user)
        elif email and not check(app.config['REGEX'], email):
            flash(f'E-mail invalide : {email}', 'error')
            user = get_user_by_id(session['login_id'])
            return render_template('new.html', form_data=request.form, user=user)
        else:
            # Create job
            job = Job(name=request.form.get('name'), url=request.form.get('url'),
                      zipCode=request.form.get('zipCode'), company=request.form.get('company'),
                      contact=request.form.get('contact'), date=datetime.now(), email=email,
                      user_id=session['login_id'])
            # Save cover letter text if provided
            cover_text = request.form.get('cover_letter_text')
            if cover_text:
                job.cover_letter_text = cover_text
            db.session.add(job)
            db.session.commit()

            # Handle file upload (capture and cover letter)
            success, error_msg = handle_file_upload(job.id)
            if not success:
                flash(error_msg, 'error')
                user = get_user_by_id(session['login_id'])
                # in case of error, show new form with previous values
                return render_template('new.html', form_data=request.form, user=user)

            # set flags
            if success and 'capture_file' in request.files and request.files['capture_file'].filename:
                job.is_capture = 1
            db.session.commit()

            flash('Record was successfully added')
            return redirect(url_for('show_all'))
    else:
        user = get_user_by_id(session['login_id'])
        return render_template('new.html', form_data=None, user=user)


@app.route('/toggle_expired/<int:id>', methods=['POST'])
@is_connected
@is_editor_or_admin
def toggle_expired(id):
    job = Job.query.get_or_404(id)
    job.active = not job.active
    db.session.commit()
    return '', 200


@app.route('/delete/<int:id>', methods=['GET', 'POST'])
@is_connected
@is_editor_or_admin
def delete(id):
    app.logger.debug(f'Delete job #{id}')
    if request.method == 'GET':
        job = Job.query.get_or_404(id)
        app.logger.debug(f'Job debug: {job}')
        db.session.delete(job)
        db.session.commit()
        flash(f'Job offer \"{job.name}\" from  \"{job.contact}\" was deleted!')
        return redirect(url_for('show_all'))


@app.route('/delete_account/<int:id>', methods=['GET', 'POST'])
@is_connected
@is_admin
def delete_account(id):
    app.logger.debug(f'Delete user #{id}')
    if request.method == 'GET':
        user = User.query.get_or_404(id)
        app.logger.debug(f'User debug: {user}')
        db.session.delete(user)
        db.session.commit()
        flash(f'User \"{user.username}\" has been deleted!')
        return redirect(url_for('show_accounts'))


@app.route('/update_account/<int:id>', methods=['GET', 'POST'])
@is_connected
def update_account(id):
    """Permet à un administrateur de changer le rôle d'un utilisateur,
    et à un utilisateur propriétaire de modifier ses propres options (allow_view_offers)
    et d'uploader son CV JSON. Les administrateurs NE PEUVENT PAS modifier
    l'option de partage `allow_view_offers` pour d'autres comptes.
    """
    user: User = User.query.get_or_404(id)
    current = get_user_by_id(session['login_id'])

    # autorisation : soit admin, soit le propriétaire du compte
    if not (current.is_admin or current.id == user.id):
        flash('Action non autorisée.', 'error')
        return redirect(url_for('show_accounts'))

    if request.method == 'POST':
        # Rôle : uniquement modifiable par un administrateur
        if current.is_admin:
            roles = ['Administrateur', 'Utilisateur']
            selected = request.form.get('role')
            # map role label to enum value (ADMIN=0, USER=1)
            if selected == 'Administrateur':
                user.role = 0
            else:
                user.role = 1

        # L'option allow_view_offers ne peut être changée que par le propriétaire
        if current.id == user.id:
            user.allow_view_offers = True if request.form.get('allow_view_offers') in ('on', 'true', '1') else False

        # Handle CV JSON upload (file input name 'cv_json') - only owner can upload
        if current.id == user.id and 'cv_json' in request.files and request.files['cv_json'].filename:
            f = request.files['cv_json']
            # accept only JSON mime/extension
            if not f.filename.lower().endswith('.json'):
                flash('Veuillez fournir un fichier JSON avec l’extension .json.', 'error')
                return redirect(url_for('update_account', id=user.id))
            # Read and validate JSON content before saving
            try:
                raw = f.read()
                import json as _json
                parsed = _json.loads(raw)
            except Exception as exc:
                app.logger.debug(f'Invalid CV JSON upload for user {user.id}: {exc}')
                flash('Fichier JSON invalide. Vérifiez le contenu et réessayez.', 'error')
                return redirect(url_for('update_account', id=user.id))
            # Persist pretty-printed JSON into user's directory
            user_dir = Path(app.static_folder) / 'uploads' / 'users' / str(user.id)
            user_dir.mkdir(parents=True, exist_ok=True)
            orig_path = user_dir / 'cv_original.json'
            with open(orig_path, 'w', encoding='utf-8') as fh:
                fh.write(_json.dumps(parsed, ensure_ascii=False, indent=2))

        # Handle PDF upload from account page (input name 'cv_pdf')
        if current.id == user.id and 'cv_pdf' in request.files and request.files['cv_pdf'].filename:
            fpdf = request.files['cv_pdf']
            if not fpdf.filename.lower().endswith('.pdf'):
                flash('Veuillez fournir un fichier PDF.', 'error')
                return redirect(url_for('update_account', id=user.id))
            cvs_dir = Path(app.static_folder) / 'uploads' / 'users' / str(user.id) / 'cvs'
            cvs_dir.mkdir(parents=True, exist_ok=True)
            dest = cvs_dir / fpdf.filename
            try:
                fpdf.save(str(dest))
            except Exception as exc:
                flash(f"Erreur lors de l'enregistrement du PDF : {exc}", 'error')
                return redirect(url_for('update_account', id=user.id))
            # update metadata
            meta_path = cvs_dir / 'cvs_meta.json'
            try:
                meta = _json.loads(meta_path.read_text(encoding='utf-8')) if meta_path.exists() else {}
            except Exception:
                meta = {}
            meta.setdefault(fpdf.filename, fpdf.filename)
            if '_default' not in meta:
                meta['_default'] = fpdf.filename
            meta_path.write_text(_json.dumps(meta, ensure_ascii=False, indent=2), encoding='utf-8')
        # Handle profile photo upload (input name 'profile_photo')
        if current.id == user.id and 'profile_photo' in request.files and request.files['profile_photo'].filename:
            pf = request.files['profile_photo']
            # accept common image extensions
            if not any(pf.filename.lower().endswith(ext) for ext in ('.png', '.jpg', '.jpeg')):
                flash('Veuillez fournir une image (png/jpg).', 'error')
                return redirect(url_for('update_account', id=user.id))
            photo_dir = Path(app.static_folder) / 'uploads' / 'users' / str(user.id)
            photo_dir.mkdir(parents=True, exist_ok=True)
            dest_photo = photo_dir / 'photo.jpg'
            try:
                pf.save(str(dest_photo))
                flash('Photo de profil enregistrée.', 'success')
            except Exception as exc:
                app.logger.error(f'Failed to save profile photo for user {user.id}: {exc}')
                flash(f"Erreur lors de l'enregistrement de la photo : {exc}", 'error')
                return redirect(url_for('update_account', id=user.id))

        db.session.add(user)
        db.session.commit()
        flash('Record was successfully updated')
        # redirect to accounts for admins, to profile for owner
        if current.is_admin and current.id != user.id:
            return redirect(url_for('show_accounts'))
        return redirect(url_for('welcome'))

    else:
        # Pass current user id so template can hide/share controls
        # Also provide list of uploaded CVs and metadata so the template can offer a "CV par défaut" selector
        cvs_dir = Path(app.static_folder) / 'uploads' / 'users' / str(user.id) / 'cvs'
        import json as _json
        meta = {}
        files = []
        if cvs_dir.exists():
            meta_path = cvs_dir / 'cvs_meta.json'
            try:
                meta = _json.loads(meta_path.read_text(encoding='utf-8')) if meta_path.exists() else {}
            except Exception:
                meta = {}
            for p in sorted(cvs_dir.iterdir()):
                if p.is_file():
                    files.append({'name': p.name, 'label': meta.get(p.name, p.name)})

        return render_template('update_account.html', user=user, current_user_id=current.id,
                               cvs_files=files, cvs_meta=meta)


@app.route('/import_cv', methods=['POST'])
@is_connected
def import_cv():
    """Endpoint pour importer un CV via upload JSON ou via URL (LinkedIn/FT).
    - Form fields possibles: 'cv_json' file, 'cv_url' text, 'user_id' hidden.
    """
    current = get_user_by_id(session['login_id'])
    target_id = int(request.form.get('user_id', current.id))
    if target_id != current.id and not current.is_admin:
        flash('Action non autorisée.', 'error')
        return redirect(url_for('update_account', id=target_id))

    import json as _json

    # Handle JSON upload
    if 'cv_json' in request.files and request.files['cv_json'].filename:
        f = request.files['cv_json']
        raw = f.read()
        try:
            parsed = parse_json_cv(raw)
        except Exception as exc:
            # If AJAX, return JSON error
            if request.is_json or request.headers.get('X-Requested-With') == 'XMLHttpRequest' or request.headers.get('Accept','').startswith('application/json'):
                return jsonify({'ok': False, 'message': f'JSON invalide: {exc}'}), 400
            flash(f'JSON invalide: {exc}', 'error')
            return redirect(url_for('update_account', id=target_id))
        # Save pretty JSON
        user_dir = Path(app.static_folder) / 'uploads' / 'users' / str(target_id)
        user_dir.mkdir(parents=True, exist_ok=True)
        orig_path = user_dir / 'cv_original.json'
        with open(orig_path, 'w', encoding='utf-8') as fh:
            import json as _json
            fh.write(_json.dumps(parsed, ensure_ascii=False, indent=2))
        # Respond differently for AJAX vs normal form
        if request.is_json or request.headers.get('X-Requested-With') == 'XMLHttpRequest' or request.headers.get('Accept','').startswith('application/json'):
            return jsonify({'ok': True, 'message': 'CV JSON importé avec succès.'}), 200
        flash('CV JSON importé avec succès.', 'success')
        return redirect(url_for('update_account', id=target_id))

    # Handle PDF upload (LinkedIn exported PDF)
    if 'cv_pdf' in request.files and request.files['cv_pdf'].filename:
        f = request.files['cv_pdf']
        filename = f.filename
        # ensure directory
        cvs_dir = Path(app.static_folder) / 'uploads' / 'users' / str(target_id) / 'cvs'
        cvs_dir.mkdir(parents=True, exist_ok=True)
        dest = cvs_dir / filename
        try:
            f.save(str(dest))
        except Exception as exc:
            flash(f"Erreur lors de l'enregistrement du PDF : {exc}", 'error')
            return redirect(url_for('update_account', id=target_id))
        # Try to parse LinkedIn PDF into structured JSON using tools.cv_tools
        parse_debug_msgs = []
        try:
            from tools.cv_tools import parse_linkedin_pdf_to_json
            parsed_out = cvs_dir / 'cv_linkedin_parsed.json'
            msg = f'Attempting parse_linkedin_pdf_to_json for {dest} -> {parsed_out}'
            app.logger.info(msg)
            parse_debug_msgs.append(msg)
            try:
                parsed_ok = parse_linkedin_pdf_to_json(dest, parsed_out)
            except Exception as e:
                app.logger.exception(f'parse_linkedin_pdf_to_json raised: {e}')
                parse_debug_msgs.append(f'parse_linkedin_pdf_to_json raised: {e}')
                parsed_ok = False

            # Fallback: if parser returned False or did not create file, try the bundled linkedin-resume-parser CLI
            if not parsed_ok or not parsed_out.exists():
                try:
                    import subprocess, sys
                    msg = 'Primary parser failed or produced no output — trying linkedin-resume-parser CLI fallback'
                    app.logger.info(msg)
                    parse_debug_msgs.append(msg)
                    # Call the CLI module provided in linkedin-resume-parser package
                    cmd = [sys.executable, '-m', 'linkedin_resume_parser.cli', str(dest), '-o', str(parsed_out)]
                    app.logger.debug(f'Running fallback command: {cmd}')
                    parse_debug_msgs.append(f'Running fallback command: {cmd}')
                    proc = subprocess.run(cmd, stdout=subprocess.PIPE, stderr=subprocess.PIPE, timeout=30)
                    app.logger.debug(f'fallback stdout: {proc.stdout[:300]}')
                    app.logger.debug(f'fallback stderr: {proc.stderr[:300]}')
                    parse_debug_msgs.append(f'fallback stdout: {proc.stdout[:300]}')
                    parse_debug_msgs.append(f'fallback stderr: {proc.stderr[:300]}')
                    parsed_ok = parsed_out.exists()
                    # If subprocess fallback failed due to module not found (module not installed in venv),
                    # try an in-process fallback using the local package present in the workspace.
                    if not parsed_ok and b"ModuleNotFoundError" in (proc.stderr or b""):
                        try:
                            app.logger.info('Subprocess fallback reported ModuleNotFoundError — attempting in-process parser import')
                            parse_debug_msgs.append('Subprocess fallback ModuleNotFoundError — trying in-process import')
                            # Attempt to import the local workspace package by adding its parent dir to sys.path
                            import importlib, sys
                            pkg_path = Path(__file__).parent / 'linkedin-resume-parser'
                            if str(pkg_path) not in sys.path:
                                sys.path.insert(0, str(pkg_path))
                            # Now import the parser module
                            lrp = importlib.import_module('linkedin_resume_parser.parser')
                            resume_dict = lrp.parse_pdf(str(dest))
                            # write JSON
                            parsed_out.write_text(_json.dumps(resume_dict, ensure_ascii=False, indent=2), encoding='utf-8')
                            parsed_ok = parsed_out.exists()
                            parse_debug_msgs.append('In-process parser produced output' if parsed_ok else 'In-process parser produced no output')
                        except Exception as e:
                            app.logger.exception(f'In-process linkedin parser failed: {e}')
                            parse_debug_msgs.append(f'In-process linkedin parser failed: {e}')
                            parsed_ok = False
                except Exception as e:
                    app.logger.exception(f'Fallback CLI parse failed: {e}')
                    parse_debug_msgs.append(f'Fallback CLI parse failed: {e}')
                    parsed_ok = False

            if parsed_ok and parsed_out.exists():
                parsed_name = parsed_out.name
            else:
                parsed_name = None
        except Exception:
            app.logger.exception('Unexpected error while attempting LinkedIn PDF parse')
            parse_debug_msgs.append('Unexpected error while attempting LinkedIn PDF parse')
            parsed_name = None

        # update metadata file
        meta_path = cvs_dir / 'cvs_meta.json'
        try:
            meta = _json.loads(meta_path.read_text(encoding='utf-8')) if meta_path.exists() else {}
        except Exception:
            meta = {}
        meta.setdefault(filename, filename)
        # register parsed json if available
        if parsed_name:
            meta.setdefault(parsed_name, parsed_name)
        if '_default' not in meta:
            # prefer parsed JSON as default, else the uploaded PDF
            meta['_default'] = parsed_name or filename
        meta_path.write_text(_json.dumps(meta, ensure_ascii=False, indent=2), encoding='utf-8')
        # If this was an AJAX/fetch request, return JSON so the frontend can handle it
        if request.is_json or request.headers.get('X-Requested-With') == 'XMLHttpRequest' or request.headers.get('Accept','').startswith('application/json'):
            return jsonify({'ok': True, 'message': 'CV PDF importé avec succès.', 'pdf': filename, 'parsed_json': parsed_name, 'parse_log': parse_debug_msgs}), 200
        flash('CV PDF importé avec succès.', 'success')
        return redirect(url_for('update_account', id=target_id))

    # Handle URL import
    cv_url = request.form.get('cv_url', '').strip()
    if cv_url:
        # basic source detection
        try:
            if 'linkedin.com' in cv_url:
                cv = fetch_linkedin_profile(cv_url)
            else:
                # try to fetch raw JSON
                import requests
                r = requests.get(cv_url, timeout=5)
                r.raise_for_status()
                cv = parse_json_cv(r.content)
        except Exception as exc:
            flash(f'Import via URL échoué : {exc}', 'error')
            return redirect(url_for('update_account', id=target_id))
        # persist minimal cv dict as JSON
        user_dir = Path(app.static_folder) / 'uploads' / 'users' / str(target_id)
        user_dir.mkdir(parents=True, exist_ok=True)
        orig_path = user_dir / 'cv_original.json'
        with open(orig_path, 'w', encoding='utf-8') as fh:
            import json as _json
            fh.write(_json.dumps(cv, ensure_ascii=False, indent=2))
        flash('CV importé depuis URL avec succès.', 'success')
        return redirect(url_for('update_account', id=target_id))

    flash('Aucun fichier ni URL fourni pour l\'import.', 'error')
    return redirect(url_for('update_account', id=target_id))


@app.route('/manage_cvs/<int:user_id>', methods=['POST'])
@is_connected
def manage_cvs(user_id):
    """Rename, set default or delete uploaded CVs (PDF or JSON)."""
    current = get_user_by_id(session['login_id'])
    if user_id != current.id and not current.is_admin:
        flash('Action non autorisée.', 'error')
        return redirect(url_for('update_account', id=user_id))

    user_dir = Path(app.static_folder) / 'uploads' / 'users' / str(user_id) / 'cvs'
    meta_path = user_dir / 'cvs_meta.json'
    if not user_dir.exists():
        flash('Aucun CV trouvé.', 'error')
        return redirect(url_for('update_account', id=user_id))

    # load metadata
    import json as _json
    try:
        meta = _json.loads(meta_path.read_text(encoding='utf-8')) if meta_path.exists() else {}
    except Exception:
        meta = {}

    # actions: rename, set_default, delete
    action = request.form.get('action')
    file = request.form.get('file')
    if not file:
        flash('Fichier non spécifié', 'error')
        return redirect(url_for('update_account', id=user_id))

    file_path = user_dir / file
    if action == 'delete':
        try:
            if file_path.exists():
                file_path.unlink()
            if file in meta:
                del meta[file]
            # if default was this, remove
            if meta.get('_default') == file:
                meta.pop('_default', None)
            meta_path.write_text(_json.dumps(meta, ensure_ascii=False, indent=2), encoding='utf-8')
            flash('Fichier supprimé', 'success')
        except Exception as e:
            flash(f'Erreur suppression: {e}', 'error')
    elif action == 'set_default':
        meta['_default'] = file
        meta_path.write_text(_json.dumps(meta, ensure_ascii=False, indent=2), encoding='utf-8')
        flash('CV par défaut mis à jour', 'success')
    elif action == 'rename':
        newname = request.form.get('newname', '').strip()
        if not newname:
            flash('Nouveau nom vide', 'error')
        else:
            meta[file] = newname
            meta_path.write_text(_json.dumps(meta, ensure_ascii=False, indent=2), encoding='utf-8')
            flash('Nom mis à jour', 'success')
    else:
        flash('Action inconnue', 'error')

    return redirect(url_for('update_account', id=user_id))


@app.route('/edit_cv/<int:user_id>', methods=['GET', 'POST'])
@is_connected
def edit_cv(user_id):
    current = get_user_by_id(session['login_id'])
    if user_id != current.id and not current.is_admin:
        flash('Action non autorisée.', 'error')
        return redirect(url_for('show_all'))
    user_dir = Path(app.static_folder) / 'uploads' / 'users' / str(user_id)
    orig_path = user_dir / 'cv_original.json'
    if request.method == 'POST':
        content = request.form.get('cv_text', '').strip()
        try:
            import json as _json
            parsed = _json.loads(content)
        except Exception as exc:
            flash(f'JSON invalide : {exc}', 'error')
            return render_template('edit_cv.html', user_id=user_id, cv_text=content)
        user_dir.mkdir(parents=True, exist_ok=True)
        with open(orig_path, 'w', encoding='utf-8') as fh:
            fh.write(_json.dumps(parsed, ensure_ascii=False, indent=2))
        flash('CV sauvegardé.', 'success')
        return redirect(url_for('update_account', id=user_id))
    else:
        if orig_path.exists():
            text = orig_path.read_text(encoding='utf-8')
        else:
            text = ''
        return render_template('edit_cv.html', user_id=user_id, cv_text=text)


@app.route('/manage_cvs_view/<int:user_id>', methods=['GET'])
@is_connected
def manage_cvs_view(user_id):
    """Affiche l'interface de gestion des CV (PDF + JSON) et propose la fusion."""
    current = get_user_by_id(session['login_id'])
    if user_id != current.id and not current.is_admin:
        flash('Action non autorisée.', 'error')
        return redirect(url_for('update_account', id=user_id))

    user_dir = Path(app.static_folder) / 'uploads' / 'users' / str(user_id) / 'cvs'
    if not user_dir.exists():
        flash('Aucun CV trouvé pour cet utilisateur.', 'info')
        return redirect(url_for('update_account', id=user_id))

    # load metadata
    import json as _json
    meta_path = user_dir / 'cvs_meta.json'
    try:
        meta = _json.loads(meta_path.read_text(encoding='utf-8')) if meta_path.exists() else {}
    except Exception:
        meta = {}

    files = []
    for p in sorted(user_dir.iterdir()):
        if p.is_file():
            files.append({
                'name': p.name,
                'path': f'uploads/users/{user_id}/cvs/{p.name}',
                'is_json': p.suffix.lower() == '.json',
                'label': meta.get(p.name, p.name),
            })

    # Build list of JSON files (for merging) but only originals (LinkedIn export or GitConnected)
    def _is_original_json(name: str) -> bool:
        ln = name.lower()
        if ln.startswith('cv_merged'):
            return False
        # heuristics: linkedin parsed exports, gitconnected, original resume files
        if 'linkedin' in ln or 'gitconnected' in ln or ln.endswith('_resume.json') or ln == 'cv_original.json':
            return True
        # also accept filenames that contain 'profile' + 'resume'
        if 'resume' in ln or 'profile' in ln:
            return True
        return False

    json_files = [f for f in files if f['is_json'] and _is_original_json(f['name'])]

    # Sections to propose selection for
    sections = ['basics', 'work', 'education', 'skills', 'projects', 'certificates', 'references', 'languages', 'volunteer']

    # Validate that listed json_files are valid JSON, annotate a preview snippet
    # and build detailed per-section items for granular selection in the template
    import json as _json
    json_details = {}
    for jf in list(json_files):
        try:
            p = Path(app.static_folder) / 'uploads' / 'users' / str(user_id) / 'cvs' / jf['name']
            txt = p.read_text(encoding='utf-8')
            parsed = _json.loads(txt)
            # preview: for basics or first keys
            preview = ''
            if isinstance(parsed.get('basics'), dict):
                preview = parsed.get('basics', {}).get('label') or parsed.get('basics', {}).get('name','')
            if not preview:
                for k in ('work','education','skills'):
                    v = parsed.get(k)
                    if v:
                        if isinstance(v, list) and len(v)>0:
                            first = v[0]
                            if isinstance(first, dict):
                                preview = first.get('position') or first.get('name') or str(first)
                            else:
                                preview = str(first)
                            break
            jf['preview'] = preview

            # Build per-section item lists for granular selection
            details = {}
            for sec in ('work','education','skills','projects','certificates','references'):
                items = parsed.get(sec) or []
                normalized = []
                if isinstance(items, list):
                    for i, it in enumerate(items):
                        if isinstance(it, dict):
                            label = it.get('position') or it.get('name') or it.get('institution') or it.get('title') or str(it)
                        else:
                            label = str(it)
                        normalized.append({'idx': i, 'label': label})
                details[sec] = normalized
            json_details[jf['name']] = details
        except Exception:
            json_files.remove(jf)

    return render_template('manage_cvs.html', user=get_user_by_id(session['login_id']),
                           files=files, json_files=json_files, sections=sections, user_id=user_id,
                           json_details=json_details)


@app.route('/merge_cvs_editor/<int:user_id>', methods=['GET', 'POST'])
@is_connected
def merge_cvs_editor(user_id):
    """Advanced merge editor: compare two source JSON side-by-side and pick items per section.
    - GET: show selector to choose left/right source (defaults to first two originals)
    - POST: receives picks and performs merge similarly to /merge_cvs
    """
    current = get_user_by_id(session['login_id'])
    if user_id != current.id and not current.is_admin:
        flash('Action non autorisée.', 'error')
        return redirect(url_for('manage_cvs_view', user_id=user_id))

    user_dir = Path(app.static_folder) / 'uploads' / 'users' / str(user_id) / 'cvs'
    if not user_dir.exists():
        flash('Aucun CV trouvé pour cet utilisateur.', 'error')
        return redirect(url_for('manage_cvs_view', user_id=user_id))

    # collect original json files
    def is_orig(n: str) -> bool:
        ln = n.lower()
        return (not ln.startswith('cv_merged')) and ('linkedin' in ln or 'gitconnected' in ln or ln.endswith('_resume.json') or ln=='cv_original.json' or 'resume' in ln or 'profile' in ln)

    json_paths = [p for p in sorted(user_dir.iterdir()) if p.is_file() and p.suffix.lower()=='.json' and is_orig(p.name)]
    json_names = [p.name for p in json_paths]

    sections = ['work','education','skills','projects','certificates','references']

    if request.method == 'GET':
        left = request.args.get('left') or (json_names[0] if len(json_names)>0 else None)
        right = request.args.get('right') or (json_names[1] if len(json_names)>1 else None)
        # load details for each
        import json as _json
        left_details = {}
        right_details = {}
        for p in json_paths:
            if p.name==left:
                left_details = _json.loads(p.read_text(encoding='utf-8'))
            if p.name==right:
                right_details = _json.loads(p.read_text(encoding='utf-8'))
        return render_template('merge_cvs_editor.html', user=get_user_by_id(session['login_id']), user_id=user_id,
                               json_names=json_names, left=left, right=right, sections=sections,
                               left_details=left_details, right_details=right_details)

    # POST: perform merge from selections
    # reuse merge_cvs behavior: accept section_* for whole-section replacement and pick__{filename}__{section}[] for granular
    selections = {}
    for key, val in request.form.items():
        if key.startswith('section_') and val:
            selections[key.replace('section_','')] = val

    # perform base merge using existing utility
    json_paths_all = [p for p in sorted(user_dir.iterdir()) if p.is_file() and p.suffix.lower()=='.json']
    try:
        merged = merge_cv_jsons(json_paths_all, selections)
    except Exception as exc:
        app.logger.error(f'merge_cvs_editor error: {exc}')
        flash(f'Erreur lors de la fusion: {exc}', 'error')
        return redirect(url_for('manage_cvs_view', user_id=user_id))

    # then apply granular picks similar to merge_cvs
    import json as _json
    for sec in sections:
        combined = []
        for p in json_paths_all:
            fname = p.name
            field = f'pick__{fname}__{sec}[]'
            picks = request.form.getlist(field)
            if picks:
                try:
                    src = _json.loads(p.read_text(encoding='utf-8'))
                    items = src.get(sec) or []
                    for idx in picks:
                        try:
                            i = int(idx)
                            if 0 <= i < len(items):
                                combined.append(items[i])
                        except Exception:
                            continue
                except Exception:
                    continue
        if combined:
            seen = set()
            uniq = []
            for it in combined:
                key = _json.dumps(it, sort_keys=True)
                if key not in seen:
                    seen.add(key)
                    uniq.append(it)
            merged[sec] = uniq

    # Save merged
    import time
    fname = f'cv_merged_editor_{int(time.time())}.json'
    dest = user_dir / fname
    try:
        dest.write_text(_json.dumps(merged, ensure_ascii=False, indent=2), encoding='utf-8')
    except Exception as exc:
        app.logger.error(f'Failed to write merged CV editor: {exc}')
        flash(f'Impossible de sauvegarder le CV fusionné: {exc}', 'error')
        return redirect(url_for('manage_cvs_view', user_id=user_id))

    # update meta
    meta_path = user_dir / 'cvs_meta.json'
    try:
        meta = _json.loads(meta_path.read_text(encoding='utf-8')) if meta_path.exists() else {}
    except Exception:
        meta = {}
    meta.setdefault(fname, fname)
    meta['_default'] = fname
    meta_path.write_text(_json.dumps(meta, ensure_ascii=False, indent=2), encoding='utf-8')

    flash(f'CV fusionné créé: {fname}', 'success')
    return redirect(url_for('manage_cvs_view', user_id=user_id))


@app.route('/merge_cvs/<int:user_id>', methods=['POST'])
@is_connected
def merge_cvs(user_id):
    """Exécute la fusion des CV JSON selon les sélections envoyées par le formulaire."""
    current = get_user_by_id(session['login_id'])
    if user_id != current.id and not current.is_admin:
        return jsonify({'error': 'Action non autorisée.'}), 403

    user_dir = Path(app.static_folder) / 'uploads' / 'users' / str(user_id) / 'cvs'
    if not user_dir.exists():
        flash('Aucun CV JSON trouvé.', 'error')
        return redirect(url_for('manage_cvs_view', user_id=user_id))

    # Collect JSON files available
    json_paths = [p for p in sorted(user_dir.iterdir()) if p.is_file() and p.suffix.lower() == '.json']
    if not json_paths:
        flash('Aucun fichier JSON à fusionner.', 'error')
        return redirect(url_for('manage_cvs_view', user_id=user_id))

    # Build selections per section from form (e.g. form fields named section_basics, section_work, ...)
    selections = {}
    for key, val in request.form.items():
        if key.startswith('section_') and val:
            section = key.replace('section_', '')
            selections[section] = val

    # Ensure selections reference allowed original files only
    def _is_original_json_local(name: str) -> bool:
        ln = name.lower()
        if ln.startswith('cv_merged'):
            return False
        if 'linkedin' in ln or 'gitconnected' in ln or ln.endswith('_resume.json') or ln == 'cv_original.json':
            return True
        if 'resume' in ln or 'profile' in ln:
            return True
        return False

    allowed = {p.name for p in json_paths if _is_original_json_local(p.name)}
    for sec, fname in list(selections.items()):
        if fname and fname not in allowed:
            flash(f'Fichier sélectionné non autorisé pour la fusion: {fname}', 'error')
            return redirect(url_for('manage_cvs_view', user_id=user_id))

    # call merge util
    try:
        merged = merge_cv_jsons(json_paths, selections)
    except Exception as exc:
        app.logger.error(f'merge_cvs error: {exc}')
        flash(f'Erreur lors de la fusion: {exc}', 'error')
        return redirect(url_for('manage_cvs_view', user_id=user_id))

    # Handle granular picks: for each sec and filename, gather picked indices
    # Expected form field names: pick__{filename}__{section}[] -> list of indices as strings
    granular_sections = ['work','education','skills','projects','certificates','references']
    import json as _json
    for sec in granular_sections:
        # collect per-file picks
        combined = []
        for p in json_paths:
            fname = p.name
            field = f'pick__{fname}__{sec}[]'
            picks = request.form.getlist(field)
            if picks:
                # load source JSON
                try:
                    src = _json.loads(p.read_text(encoding='utf-8'))
                    items = src.get(sec) or []
                    for idx in picks:
                        try:
                            i = int(idx)
                            if 0 <= i < len(items):
                                combined.append(items[i])
                        except Exception:
                            continue
                except Exception:
                    continue
        if combined:
            # deduplicate simple: keep first occurrence by stringified content
            seen = set()
            uniq = []
            for it in combined:
                key = _json.dumps(it, sort_keys=True)
                if key not in seen:
                    seen.add(key)
                    uniq.append(it)
            merged[sec] = uniq

    # Save merged as new JSON file
    import json as _json, time
    fname = f'cv_merged_{int(time.time())}.json'
    dest = user_dir / fname
    try:
        dest.write_text(_json.dumps(merged, ensure_ascii=False, indent=2), encoding='utf-8')
    except Exception as exc:
        app.logger.error(f'Failed to write merged CV: {exc}')
        flash(f'Impossible de sauvegarder le CV fusionné: {exc}', 'error')
        return redirect(url_for('manage_cvs_view', user_id=user_id))

    # Try to generate a PDF for convenience and save alongside JSON
    try:
        pdf_bytes = generate_tailored_cv_pdf_bytes(merged)
        pdf_name = fname.replace('.json', '.pdf')
        pdf_path = user_dir / pdf_name
        with open(pdf_path, 'wb') as pf:
            pf.write(pdf_bytes)
    except Exception as exc:
        app.logger.debug(f'PDF generation skipped or failed for merged CV: {exc}')

    # update metadata
    meta_path = user_dir / 'cvs_meta.json'
    try:
        meta = _json.loads(meta_path.read_text(encoding='utf-8')) if meta_path.exists() else {}
    except Exception:
        meta = {}
    meta.setdefault(fname, fname)
    # if pdf was created, set it as default preview
    if 'pdf_name' in locals():
        meta.setdefault(pdf_name, pdf_name)
        meta['_default'] = pdf_name
    else:
        meta['_default'] = fname  # set merged as default by convenience
    meta_path.write_text(_json.dumps(meta, ensure_ascii=False, indent=2), encoding='utf-8')

    flash(f'CV fusionné créé: {fname}', 'success')
    return redirect(url_for('manage_cvs_view', user_id=user_id))


@app.route('/update/<int:id>', methods=['GET', 'POST'])
@is_connected
@is_editor_or_admin
def update(id):
    job: Job = Job.query.get_or_404(id)
    def _parse_form_date(value: Optional[str], original: Optional[datetime] = None) -> Optional[datetime]:
        """Parse a date/time coming from the form.
        - Si value est vide, retourne None.
        - Si le formulaire envoie uniquement une date (YYYY-MM-DD, sans heure),
          on préserve l'heure de 'original' pour ne pas écraser le composant horaire
          déjà stocké en base.
        - Accepte aussi les formats datetime-local complets.
        """
        if not value:
            return None
        v = value.strip()
        # Formats avec heure : on les utilise tels quels
        for fmt in ("%Y-%m-%dT%H:%M:%S", "%Y-%m-%dT%H:%M", "%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M"):
            try:
                return datetime.strptime(v, fmt)
            except Exception:
                continue
        # Format date seule (YYYY-MM-DD) : préserver l'heure originale si disponible
        try:
            parsed_date = datetime.strptime(v, "%Y-%m-%d")
            if original is not None:
                return parsed_date.replace(hour=original.hour, minute=original.minute,
                                           second=original.second, microsecond=original.microsecond)
            return parsed_date
        except Exception:
            pass
        # Dernier recours : fromisoformat
        try:
            return datetime.fromisoformat(v)
        except Exception:
            pass
        return None

    if request.method == 'POST':
        job.name = request.form.get('name')
        job.url = request.form.get('url')
        job.zipCode = request.form.get('zipCode')
        job.company = request.form.get('company')
        job.contact = request.form.get('contact')
        job.email = request.form.get('email')

        application_date: str = request.form.get('applicationDate')
        app.logger.debug(f'application_date raw: {application_date}')
        job.applicationDate = _parse_form_date(application_date, original=job.applicationDate)

        relaunch_date: str = request.form.get('relaunchDate')
        job.relaunchDate = _parse_form_date(relaunch_date, original=job.relaunchDate)

        refusal_date: str = request.form.get('refusalDate')
        job.refusalDate = _parse_form_date(refusal_date, original=job.refusalDate)

        # cover letter text
        cover_text = request.form.get('cover_letter_text')
        job.cover_letter_text = cover_text if cover_text else None

        # Handle file upload
        success, error_msg = handle_file_upload(job.id)
        if not success:
            flash(error_msg, 'error')
            return render_template('update.html', job=job)

        if success and 'capture_file' in request.files and request.files['capture_file'].filename:
            job.is_capture = 1

        db.session.commit()
        flash('Record was successfully updated')
        return redirect(url_for('show_all'))
    else:
        return render_template('update.html', job=job)


@app.before_request
def create_tables():
    db.create_all()
    # Migration douce : ajoute les colonnes ajoutées post-création si elles n'existent pas encore
    from sqlalchemy import text as _text
    with db.engine.connect() as _conn:
        _cols = [row[1] for row in _conn.execute(_text("PRAGMA table_info(job)"))]
        if 'ft_offer_id' not in _cols:
            _conn.execute(_text("ALTER TABLE job ADD COLUMN ft_offer_id VARCHAR(30)"))
            _conn.commit()
        # Ajouter la nouvelle colonne opt-in allow_view_offers sur la table user si absent
        _user_cols = [row[1] for row in _conn.execute(_text("PRAGMA table_info(user)"))]
        if 'allow_view_offers' not in _user_cols:
            try:
                _conn.execute(_text("ALTER TABLE user ADD COLUMN allow_view_offers BOOLEAN DEFAULT 0"))
                _conn.commit()
            except Exception:
                # Certaines versions SQLite n'acceptent pas BOOLEAN, retomber sur INTEGER
                try:
                    _conn.execute(_text("ALTER TABLE user ADD COLUMN allow_view_offers INTEGER DEFAULT 0"))
                    _conn.commit()
                except Exception:
                    pass


@app.before_request
def validate_session():
    if 'login_id' in session:
        user_session = Session.query.filter_by(login_id=session['login_id'], end=None).first()
        if not user_session:
            session.clear()
            flash('Votre session a été fermée par un administrateur.', 'warning')
            return redirect(url_for('login'))

@app.before_request
def create_captures_dir():
    directory_path = f'{os.path.dirname(__file__)}/static/images'
    if not os.path.exists(directory_path):
        try:
            os.mkdir(directory_path)
            print(f"Directory {directory_path} created successfully.")
        except Exception as e:
            print(f"Error creating directory: {e}")

@app.route('/lucky')
def upload_form():
    return render_template('upload.html', result=None)

@app.route('/upload', methods=['POST'])
def upload_file():
    result = None
    result_file = None

    if 'file' not in request.files:
        result = 'Aucun fichier n\'a été téléchargé.'
    else:
        file = request.files['file']

        if file.filename == '':
            result = 'Aucun fichier sélectionné.'
        else:
            # Sauvegarder le fichier d'entrée
            file_path = os.path.join(app.config['UPLOAD_FOLDER'], file.filename)
            file.save(file_path)

            # Traitement du fichier
            sleep(5)
            result_file_path = os.path.join(app.config['UPLOAD_FOLDER'], 'result.csv')
            # result_df.to_csv(result_file_path, index=False)

            result = 'Le fichier "{}" a été traité avec succès (Test avec sleep de 5s).'.format(file.filename)
            result_file = result_file_path

    return render_template('upload.html', result=result, result_file=result_file)

@app.route('/download/<filename>')
def download_file(filename):
    return send_file(os.path.join(app.config['UPLOAD_FOLDER'], filename), as_attachment=True)

# ─────────────────────────────────────────────────────────────────────────────
# France Travail – Recherche d'offres
# ─────────────────────────────────────────────────────────────────────────────

def _ft_ats_prefix(ats_mode: bool) -> str:
    """Retourne le préfixe ATS à injecter dans le prompt IA si le mode est activé."""
    if not ats_mode:
        return ''
    return (
        "MODE ATS ACTIVÉ – Optimise ce document pour être correctement traité par les systèmes "
        "ATS (Applicant Tracking System) des recruteurs. Règles strictes : "
        "(1) Utilise EXACTEMENT les mots-clés et formulations de l'offre d'emploi. "
        "(2) Structure le contenu avec des sections et sous-sections claires (pas de colonnes multiples). "
        "(3) Chaque point d'expérience doit commencer par un verbe d'action fort "
        "(Développé, Conçu, Déployé, Optimisé, Géré, Mis en place, Piloté…). "
        "(4) Évite les tableaux complexes, graphiques et icônes dans le contenu textuel. "
        "(5) Le titre doit correspondre mot pour mot à l'intitulé du poste visé.\n\n"
    )


@app.route('/france_travail')
@is_connected
def france_travail():
    user = get_user_by_id(session['login_id'])
    last_extraction = AppSetting.get('ft_last_extraction')
    # Construire dynamiquement la liste de mots-clés ATS à partir du CV de l'utilisateur
    try:
        # Prefer the connected user's default CV when building ATS keywords for the edit page
        cv_data = _resolve_user_cv_data(session.get('login_id')) or load_cv_data(app.static_folder)
        derived = []
        # titre / label du CV
        label = (cv_data.get('basics', {}) or {}).get('label')
        if label:
            derived.append(label)
        # noms de compétences (skills)
        for s in (cv_data.get('skills', []) or [])[:8]:
            name = s.get('name') if isinstance(s, dict) else str(s)
            if name:
                derived.append(name)
        # postes récents
        for w in (cv_data.get('work', []) or [])[:6]:
            pos = (w.get('position') or w.get('summary') or '')
            if pos:
                derived.append(pos)
    except Exception:
        derived = []

    # Combiner les mots-clés dérivés et la liste statique tout en évitant les doublons
    seen = set()
    ats_list = []
    for kw in (derived + ATS_KEYWORDS_PROFILE):
        if not kw:
            continue
        k = kw.strip()
        if k and k not in seen:
            ats_list.append(k)
            seen.add(k)
        if len(ats_list) >= 16:
            break

    return render_template(
        'france_travail.html',
        user=user,
        contract_types=CONTRACT_TYPES,
        work_modes=WORK_MODES,
        departments=DEPARTMENTS,
        ats_keywords=ats_list,
        last_extraction=last_extraction,
        default_dept=app.config.get('FT_DEFAULT_DEPT', '06'),
    )


@app.route('/france_travail/edit/<int:search_id>')
@is_connected
def france_travail_edit_page(search_id):
    """Affiche le formulaire de recherche prérempli pour éditer une recherche existante.
    Le formulaire est le même que la page `france_travail`, mais les champs sont préremplis
    et la soumission POST ira vers `/france_travail/search/<id>/edit`.
    """
    user = get_user_by_id(session['login_id'])
    ft_search = FtSearch.query.get_or_404(search_id)
    if ft_search.user_id != session['login_id']:
        flash('Action non autorisée.', 'error')
        return redirect(url_for('ft_searches'))
    params = ft_search.params or {}
    # map params to prefill variables expected by the template
    prefill = {
        'prefill_mode': params.get('mode', 'manual'),
        'prefill_mots_cles': params.get('mots_cles') or '',
        'prefill_types_contrat': params.get('types_contrat') or [],
        'prefill_departement': params.get('departement') or app.config.get('FT_DEFAULT_DEPT', '06'),
        'prefill_mode_travail': params.get('mode_travail') or '',
        'prefill_entreprises_adaptees': bool(params.get('entreprises_adaptees', False)),
        'prefill_provider': params.get('provider') or 'france_travail',
        'edit_search_id': ft_search.id,
        'search_info': ft_search.search_info,
    }
    # Même logique que la page principale : construire une liste ATS dérivée du CV
    try:
        # Prefer the connected user's default CV when building ATS keywords for the edit page
        cv_data = _resolve_user_cv_data(session.get('login_id')) or load_cv_data(app.static_folder)
        derived = []
        label = (cv_data.get('basics', {}) or {}).get('label')
        if label:
            derived.append(label)
        for s in (cv_data.get('skills', []) or [])[:8]:
            name = s.get('name') if isinstance(s, dict) else str(s)
            if name:
                derived.append(name)
        for w in (cv_data.get('work', []) or [])[:6]:
            pos = (w.get('position') or w.get('summary') or '')
            if pos:
                derived.append(pos)
    except Exception:
        derived = []

    seen = set()
    ats_list = []
    for kw in (derived + ATS_KEYWORDS_PROFILE):
        if not kw:
            continue
        k = kw.strip()
        if k and k not in seen:
            ats_list.append(k)
            seen.add(k)
        if len(ats_list) >= 16:
            break

    return render_template(
        'france_travail.html',
        user=user,
        contract_types=CONTRACT_TYPES,
        work_modes=WORK_MODES,
        departments=DEPARTMENTS,
        ats_keywords=ats_list,
        last_extraction=AppSetting.get('ft_last_extraction'),
        default_dept=app.config.get('FT_DEFAULT_DEPT', '06'),
        **prefill,
    )


@app.route('/france_travail/search', methods=['POST'])
@is_connected
def france_travail_search():
    user = get_user_by_id(session['login_id'])
    client_id     = app.config.get('FT_CLIENT_ID', '')
    client_secret = app.config.get('FT_CLIENT_SECRET', '')

    if not client_id or not client_secret:
        flash(
            'Les identifiants France Travail ne sont pas configurés. '
            'Renseignez FT_CLIENT_ID et FT_CLIENT_SECRET dans config.py.',
            'error',
        )
        return redirect(url_for('france_travail'))

    mode       = request.form.get('mode', 'manual')
    departement = request.form.get('departement', '') or None

    try:
        if mode == 'auto':
            # Prefer the connected user's default CV data for automatic searches
            cv_data = _resolve_user_cv_data(user.id) or load_cv_data(app.static_folder)
            min_creation_date = None
            reset_flag = request.form.get('reset_flag') == 'on'
            if not reset_flag:
                last_str = AppSetting.get('ft_last_extraction')
                if last_str:
                    try:
                        min_creation_date = datetime.fromisoformat(last_str)
                    except Exception:
                        pass
            offers = search_auto_from_cv(
                client_id=client_id,
                client_secret=client_secret,
                cv_data=cv_data,
                departement=departement,
                min_creation_date=min_creation_date,
            )
            # Mémoriser la date d'extraction automatique
            AppSetting.set('ft_last_extraction', datetime.utcnow().isoformat())
            depuis = f" depuis le {min_creation_date.strftime('%d/%m/%Y')}" if min_creation_date else ""
            search_info = f"Mode automatique{depuis} – {len(offers)} offre(s) trouvée(s)"
            search_params = {'mode': 'auto', 'departement': departement}
        else:
            mots_cles          = request.form.get('mots_cles', '').strip() or None
            types_contrat      = request.form.getlist('types_contrat') or None
            mode_travail       = request.form.get('mode_travail', '') or None
            entreprises_adapt  = request.form.get('entreprises_adaptees') == 'on'
            # Allow selecting an alternative provider (greenhouse, lever, ashby, teamtailor)
            provider_name = request.form.get('provider', 'france_travail')
            if provider_name and provider_name != 'france_travail':
                try:
                    from tools.jobboard_providers import get_provider
                    prov = get_provider(provider_name)
                    # Map our form params into a simple q parameter for provider adapters
                    q = mots_cles or ''
                    offers = prov.search_offers(q=q, departement=departement, types_contrat=types_contrat,
                                                mode_travail=mode_travail)
                    # provider returns normalized dicts; keep as-is
                except Exception as exc:
                    app.logger.error(f'Provider {provider_name} error: {exc}')
                    offers = []
            else:
                offers = search_offers(
                    client_id=client_id,
                    client_secret=client_secret,
                    mots_cles=mots_cles,
                    types_contrat=types_contrat,
                    departement=departement,
                    mode_travail=mode_travail,
                    entreprises_adaptees=entreprises_adapt,
                )
            kw   = mots_cles or '(tous)'
            dept = DEPARTMENTS.get(departement or '', departement or 'Toute la France')
            ct   = ', '.join(types_contrat) if types_contrat else 'Tous'
            mt   = mode_travail or 'Tous'
            parts = [f"Mots-clés : \u00ab{kw}\u00bb", f"D\u00e9partement : {dept}",
                     f"Contrats : {ct}", f"Mode : {mt}"]
            if entreprises_adapt:
                parts.append("Handi-engag\u00e9s uniquement")
            search_info = "Mode manuel – " + " | ".join(parts) + f" – {len(offers)} offre(s)"
            search_params = {
                'mode': 'manual', 'mots_cles': mots_cles, 'types_contrat': types_contrat,
                'departement': departement, 'mode_travail': mode_travail,
                'entreprises_adaptees': entreprises_adapt,
            }

    except Exception as exc:
        app.logger.error(f'France Travail search error: {exc}')
        flash(f'Erreur lors de la recherche France Travail : {exc}', 'error')
        return redirect(url_for('france_travail'))

    # Sauvegarde persistante du résultat pour consultation ultérieure
    ft_search = FtSearch(
        user_id=session['login_id'],
        search_info=search_info,
        offers=offers,
        search_params=search_params,
    )
    db.session.add(ft_search)
    db.session.commit()
    return redirect(url_for('ft_search_view', search_id=ft_search.id))


@app.route('/france_travail/searches')
@is_connected
def ft_searches():
    """Liste de toutes les recherches sauvegardées de l'utilisateur."""
    user = get_user_by_id(session['login_id'])
    searches = (
        FtSearch.query
        .filter_by(user_id=session['login_id'])
        .order_by(FtSearch.created_at.desc())
        .all()
    )
    return render_template('ft_searches.html', user=user, searches=searches)


def _ft_added_ids(owner_id: int | None = None) -> set[str]:
    """Retourne l'ensemble des ft_offer_id déjà présents dans la table job.

    Si owner_id est précisé, ne retourne que les IDs ajoutés par cet utilisateur.
    """
    q = db.session.query(Job.ft_offer_id).filter(Job.ft_offer_id.isnot(None))
    if owner_id is not None:
        q = q.filter(Job.user_id == owner_id)
    rows = q.all()
    return {r[0] for r in rows}


@app.route('/france_travail/search/<int:search_id>')
@is_connected
def ft_search_view(search_id):
    """Affiche les résultats d'une recherche sauvegardée."""
    user = get_user_by_id(session['login_id'])
    ft_search = FtSearch.query.get_or_404(search_id)
    return render_template(
        'offers_candidates.html',
        user=user,
        search=ft_search,
        offers=ft_search.offers,
        search_info=ft_search.search_info,
        search_id=search_id,
        added_ids=_ft_added_ids(session.get('login_id')),
        unavailable_ids=ft_search.unavailable,
    )


@app.route('/france_travail/search/<int:search_id>.json')
@is_connected
def ft_search_view_json(search_id):
    """Retourne la représentation JSON d'une recherche (pour préremplir le modal d'édition)."""
    ft_search = FtSearch.query.get_or_404(search_id)
    if ft_search.user_id != session.get('login_id'):
        return jsonify({'error': 'Action non autorisée.'}), 403
    return jsonify({
        'id': ft_search.id,
        'search_info': ft_search.search_info,
        'params': ft_search.params,
    })


@app.route('/france_travail/search/<int:search_id>/export', endpoint='ft_search_export')
@is_connected
def ft_search_export(search_id):
    """Export table of offers for a saved France Travail search.
    Query param `format` can be 'pdf' (default) or 'csv'. Returns a downloadable file.
    """
    ft_search = FtSearch.query.get_or_404(search_id)
    offers = ft_search.offers or []
    fmt = (request.args.get('format') or 'pdf').lower()
    # search_info can contain characters unsuitable for filenames; build a safe slug
    def _safe_slug(s: str, maxlen: int = 60) -> str:
        import re
        if not s:
            return ''
        slug = re.sub(r'\s+', '_', s)
        slug = re.sub(r'[^A-Za-z0-9_\-]', '', slug)
        return slug[:maxlen]
    search_label = _safe_slug(ft_search.search_info or f'search_{search_id}')

    if fmt in ('csv', 'xlsx'):
        # Build rows in memory first
        import csv
        from io import StringIO, BytesIO as _BytesIO
        rows = []
        # header: added URL column
        header = ['Date', 'Intitulé', 'Entreprise', 'Lieu', 'Contrat', 'Salaire', 'URL']
        rows.append(header)
        for o in offers:
            date = o.get('dateCreation') or o.get('date') or ''
            title = o.get('intitule') or o.get('title') or o.get('name') or ''
            company = o.get('entreprise') or o.get('company') or ''
            location = o.get('lieu') or o.get('location') or ''
            contract = o.get('typeContrat') or o.get('typeContratLibelle') or o.get('contract') or ''
            salary = o.get('salaire') or o.get('salary') or ''
            url = o.get('url') or o.get('link') or o.get('source_url') or ''
            rows.append([date, title, company, location, contract, salary, url])

        filename_base = f'ft_search_{search_id}_{search_label}' if search_label else f'ft_search_{search_id}'

        if fmt == 'csv':
            sio = StringIO()
            writer = csv.writer(sio)
            for r in rows:
                writer.writerow(r)
            sio.seek(0)
            filename = f'{filename_base}.csv'
            return send_file(BytesIO(sio.getvalue().encode('utf-8')), mimetype='text/csv', as_attachment=True,
                             download_name=filename)

        # fmt == 'xlsx'
        try:
            from openpyxl import Workbook
            from openpyxl.utils import get_column_letter
        except Exception:
            flash('L\'export Excel nécessite la bibliothèque openpyxl. Installez-la (pip install openpyxl).', 'error')
            return redirect(url_for('ft_search_view', search_id=search_id))

        wb = Workbook()
        ws = wb.active
        ws.title = 'Offres'
        from openpyxl.styles import Font
        link_font = Font(color='0000EE', underline='single')
        for r_idx, r in enumerate(rows, start=1):
            for c_idx, val in enumerate(r, start=1):
                cell = ws.cell(row=r_idx, column=c_idx, value=val)
                # If this is the header row, keep default styling
                if r_idx == 1:
                    continue
                # Make the title (column 2) a hyperlink to the offer URL if available (column 7)
                if c_idx == 2:
                    url_val = r[6] if len(r) > 6 else ''
                    if url_val:
                        cell.hyperlink = url_val
                        cell.font = link_font
                # Also make the URL column itself a clickable hyperlink
                if c_idx == 7 and val:
                    cell.hyperlink = val
                    cell.font = link_font
        # Auto-width for a few columns
        for i, _ in enumerate(rows[0], start=1):
            col = get_column_letter(i)
            ws.column_dimensions[col].width = 20

        bio = _BytesIO()
        wb.save(bio)
        bio.seek(0)
        filename = f'{filename_base}.xlsx'
        return send_file(bio, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                         as_attachment=True, download_name=filename)

    # default: PDF
    try:
        from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet
        from reportlab.lib.pagesizes import A4
    except Exception:
        flash('La génération PDF nécessite la bibliothèque reportlab.', 'error')
        return redirect(url_for('ft_search_view', search_id=search_id))

    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    styles = getSampleStyleSheet()
    story = []
    # Include search criteria in PDF title for clarity
    pdf_title = f"France Travail — Recherche #{search_id}"
    if ft_search.search_info:
        pdf_title += f" — {ft_search.search_info}"
    story.append(Paragraph(pdf_title, styles['Title']))
    story.append(Spacer(1, 12))
    for i, o in enumerate(offers, start=1):
        title = o.get('intitule') or o.get('title') or o.get('name') or '—'
        company = o.get('entreprise') or o.get('company') or ''
        location = o.get('lieu') or o.get('location') or ''
        contract = o.get('typeContrat') or o.get('typeContratLibelle') or ''
        salary = o.get('salaire') or o.get('salary') or ''
        url_offer = o.get('url') or o.get('link') or o.get('source_url') or ''
        # Make the title itself a clickable link in the PDF when URL is present
        if url_offer:
            title_html = f'<link href="{url_offer}"><b>{title}</b></link>'
        else:
            title_html = f'<b>{title}</b>'
        p_text = f"{i}. {title_html} — {company} — {location} — {contract} — {salary}"
        story.append(Paragraph(p_text, styles['Normal']))
        # also include the raw URL below for clarity (kept as before)
        if url_offer:
            link_p = Paragraph(f'<link href="{url_offer}">{url_offer}</link>', styles['Normal'])
            story.append(link_p)
        story.append(Spacer(1, 8))

    doc.build(story)
    buffer.seek(0)
    return send_file(buffer, mimetype='application/pdf', as_attachment=True, download_name=f'ft_search_{search_id}.pdf')


@app.route('/france_travail/search/<int:search_id>/refresh', methods=['POST'])
@is_connected
def ft_search_refresh(search_id):
    """Réactualise une recherche sauvegardée en réexécutant les mêmes paramètres."""
    ft_search = FtSearch.query.get_or_404(search_id)
    client_id     = app.config.get('FT_CLIENT_ID', '')
    client_secret = app.config.get('FT_CLIENT_SECRET', '')
    if not client_id or not client_secret:
        flash('Identifiants France Travail non configurés.', 'error')
        return redirect(url_for('ft_search_view', search_id=search_id))

    p = ft_search.params
    try:
        if p.get('mode') == 'auto':
            cv_data = _resolve_user_cv_data(session.get('login_id')) or load_cv_data(app.static_folder)
            offers = search_auto_from_cv(
                client_id=client_id, client_secret=client_secret,
                cv_data=cv_data, departement=p.get('departement'),
            )
            search_info = f"Mode automatique (actualisé) – {len(offers)} offre(s)"
        else:
            offers = search_offers(
                client_id=client_id, client_secret=client_secret,
                mots_cles=p.get('mots_cles'), types_contrat=p.get('types_contrat'),
                departement=p.get('departement'), mode_travail=p.get('mode_travail'),
                entreprises_adaptees=bool(p.get('entreprises_adaptees', False)),
            )
            kw = p.get('mots_cles') or '(tous)'
            search_info = f"Mode manuel – Mots-clés : «{kw}» – {len(offers)} offre(s) (actualisé)"
    except Exception as exc:
        app.logger.error(f'FT refresh error: {exc}')
        flash(f'Erreur lors de la réactualisation : {exc}', 'error')
        return redirect(url_for('ft_search_view', search_id=search_id))

    ft_search.offers      = offers
    ft_search.search_info = search_info
    ft_search.created_at  = datetime.utcnow()
    ft_search.unavailable_ids = '[]'
    db.session.commit()
    flash(f'Résultats actualisés – {len(offers)} offre(s) trouvée(s).', 'success')
    return redirect(url_for('ft_search_view', search_id=search_id))


@app.route('/france_travail/search/<int:search_id>/toggle_unavailable', methods=['POST'])
@is_connected
def ft_toggle_unavailable(search_id):
    """Bascule le statut 'indisponible' d'une offre dans une recherche sauvegardée."""
    ft_search = FtSearch.query.get_or_404(search_id)
    offer_id = request.form.get('offer_id', '').strip()
    if offer_id:
        ft_search.toggle_unavailable(offer_id)
        db.session.commit()
    return redirect(url_for('ft_search_view', search_id=search_id))


@app.route('/france_travail/search/<int:search_id>/add_candidature', methods=['POST'])
@is_connected
@is_editor_or_admin
def ft_add_candidature(search_id):
    """Ajoute une offre France Travail comme candidature et reste sur la page de résultats."""
    import re as _re
    offer_id  = request.form.get('offer_id', '').strip()
    title     = request.form.get('title', '').strip()
    url_offer = request.form.get('url', '').strip()
    company   = request.form.get('company', '').strip()
    location  = request.form.get('location', '').strip()

    # Déduplication par identifiant FT (plus fiable que l'URL)
    if offer_id:
        existing = Job.query.filter_by(ft_offer_id=offer_id).first()
        if existing:
            flash(f'La candidature « {title} » ({company}) a déjà été ajoutée !', 'warning')
            return redirect(url_for('ft_search_view', search_id=search_id))

    zip_code = ''
    if location:
        m = _re.search(r'\b(\d{5})\b', location)
        if m:
            zip_code = m.group(1)
        else:
            m2 = _re.search(r'\b(\d{2})\b', location)
            if m2:
                zip_code = m2.group(1)

    job = Job(
        name=title, url=url_offer, zipCode=zip_code, company=company,
        contact='', date=datetime.now(), email='', user_id=session['login_id'],
    )
    # Defensive logging: ensure owner is the logged-in user
    try:
        owner_id = int(session.get('login_id'))
    except Exception:
        owner_id = None
    app.logger.debug(f"ft_add_candidature: session_login_id={session.get('login_id')} owner_id={owner_id} search_id={search_id} offer_id={offer_id}")
    if owner_id is None:
        flash('Impossible de déterminer le propriétaire de la candidature.', 'error')
        return redirect(url_for('ft_search_view', search_id=search_id))
    job.user_id = owner_id
    job.ft_offer_id = offer_id or None
    db.session.add(job)
    db.session.commit()
    flash(f'Candidature « {title} » ({company}) ajoutée avec succès !', 'success')
    return redirect(url_for('ft_search_view', search_id=search_id))


@app.route('/france_travail/search/<int:search_id>/delete', methods=['POST'])
@is_connected
def ft_search_delete(search_id):
    """Supprime une recherche sauvegardée."""
    ft_search = FtSearch.query.get_or_404(search_id)
    if ft_search.user_id != session['login_id']:
        flash('Action non autorisée.', 'error')
        return redirect(url_for('ft_searches'))
    db.session.delete(ft_search)
    db.session.commit()
    flash('Recherche supprimée.', 'success')
    return redirect(url_for('ft_searches'))


@app.route('/france_travail/searches/delete', methods=['POST'])
@is_connected
def ft_searches_delete():
    """Supprime plusieurs recherches en masse. POST JSON: {ids: [1,2,3]}"""
    try:
        payload = request.get_json(force=True)
    except Exception:
        return jsonify({'error': 'Invalid JSON payload'}), 400
    ids = payload.get('ids') if isinstance(payload, dict) else None
    if not ids or not isinstance(ids, list):
        return jsonify({'error': 'Missing ids list'}), 400
    # ensure ownership and delete
    deleted = 0
    for sid in ids:
        try:
            sid_int = int(sid)
        except Exception:
            continue
        ft_search = FtSearch.query.get(sid_int)
        if not ft_search:
            continue
        if ft_search.user_id != session.get('login_id'):
            continue
        db.session.delete(ft_search)
        deleted += 1
    db.session.commit()
    return jsonify({'deleted': deleted}), 200


@app.route('/france_travail/search/<int:search_id>/edit', methods=['POST'])
@is_connected
def ft_search_edit(search_id):
    """Édite les attributs d'une recherche sauvegardée. Accepts JSON {search_info, search_params} or form data."""
    ft_search = FtSearch.query.get_or_404(search_id)
    if ft_search.user_id != session.get('login_id'):
        return jsonify({'error': 'Action non autorisée.'}), 403
    # Accept JSON or form
    data = {}
    try:
        data = request.get_json(silent=True) or {}
    except Exception:
        data = {}
    # fallback to form
    if not data:
        # Build structured params from form fields (same fields as france_travail form)
        form = request.form
        data = {}
        data['search_info'] = form.get('search_info')
        # build search_params dict
        sp = {}
        sp['mode'] = form.get('mode') or 'manual'
        sp['mots_cles'] = form.get('mots_cles') or None
        sp['types_contrat'] = form.getlist('types_contrat') or None
        sp['departement'] = form.get('departement') or None
        sp['mode_travail'] = form.get('mode_travail') or None
        sp['entreprises_adaptees'] = True if form.get('entreprises_adaptees') in ('on', 'true', '1') else False
        sp['provider'] = form.get('provider') or 'france_travail'
        data['search_params'] = sp
    search_info = data.get('search_info')
    search_params = data.get('search_params') or data.get('search_params_json')
    if search_info is not None:
        ft_search.search_info = search_info
    if search_params is not None:
        # accept dict or JSON string
        if isinstance(search_params, dict):
            ft_search.search_params_json = _json.dumps(search_params, ensure_ascii=False)
        else:
            try:
                parsed = _json.loads(search_params)
                ft_search.search_params_json = _json.dumps(parsed, ensure_ascii=False)
            except Exception:
                return jsonify({'error': 'Paramètres JSON invalides'}), 400
    db.session.commit()
    # If this was an AJAX/JSON request, return JSON as before
    if request.is_json or request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        return jsonify({'ok': True, 'id': ft_search.id}), 200
    # Otherwise assume a normal form POST and redirect to the search view with a flash
    flash('Recherche mise à jour.', 'success')
    return redirect(url_for('ft_search_view', search_id=ft_search.id))


@app.route('/providers/search', methods=['POST'])
@is_connected
def providers_search():
    """Endpoint utilitaire pour tester un provider externe sans passer par FT.
    POST JSON: { provider: 'teamtailor', base_url: 'https://company.teamtailor.com', api_key: '...', q: 'dev' }
    Retourne JSON listé d'offres normalisées ou l'erreur.
    """
    # Allow a development bypass header X-DEV-KEY so we can call this endpoint from curl
    dev_key = request.headers.get('X-DEV-KEY') or request.args.get('dev_key')
    allowed_dev = app.config.get('DEV_SEARCH_KEY') or os.environ.get('DEV_SEARCH_KEY') or 'devtest'
    if dev_key == allowed_dev:
        # bypass authentication
        pass
    else:
        try:
            payload = request.get_json(force=True)
        except Exception:
            return jsonify({'error': 'Invalid JSON payload'}), 400
    try:
        # if payload not defined (bypass), parse it now
        if 'payload' not in locals():
            payload = request.get_json(force=True) or {}
        payload = request.get_json(force=True)
    except Exception:
        return jsonify({'error': 'Invalid JSON payload'}), 400
    provider = (payload.get('provider') or '').strip()
    if not provider:
        return jsonify({'error': 'Missing provider'}), 400
    base_url = payload.get('base_url') or None
    api_key = payload.get('api_key') or None
    q = payload.get('q') or payload.get('mots_cles') or None
    try:
        from tools.jobboard_providers import get_provider
        prov = get_provider(provider, api_key=api_key, base_url=base_url)
        offers = prov.search_offers(q=q)
        return jsonify({'provider': provider, 'count': len(offers), 'offers': offers})
    except Exception as exc:
        app.logger.error(f'providers_search error: {exc}')
        return jsonify({'error': str(exc)}), 500


# ─────────────────────────────────────────────────────────────────────────────


@app.template_filter('format_paris_time')
def format_paris_time(utc_dt):
    paris_tz = timezone('Europe/Paris')
    paris_time = utc_dt.astimezone(paris_tz)
    return paris_time.strftime('%A %d %B %Y à %Hh%M')


# if app.config['SCHEDULER']:
#     app.logger.debug(app.config['SCHEDULER'])
#     scheduler = BackgroundScheduler()
#     scheduler.add_job(func=send_reminders, trigger="interval", seconds=app.config['SCHEDULER_INTERVAL'])
#     scheduler.start()
#
#     # Shut down the scheduler when exiting the app
#     atexit.register(lambda: scheduler.shutdown())

if __name__ == '__main__':
    app.run()
    toolbar.init_app(app)
    app.run(debug=True, use_debugger=True, use_reloader=False)
